#!/usr/bin/env python3
"""build_parameter_workbook.py — assemble the platform parameter register (XLSX).

Collects every quantitative claim in docs/00–50 into one traceable, working
workbook: a filterable parameter register plus live calculation sheets that
re-derive the headline numbers from first principles (Magnus psychrometrics and
the doc 00 §4 airflow–moisture model), so the published figures can be checked
rather than trusted.

Usage:  python3 scripts/build_parameter_workbook.py [output.xlsx]
Deps:   openpyxl >= 3.1

Conventions (engineering-register practice):
  * Every row carries unit, confidence grade, gating test, derivation, source.
  * Ranges are held as MIN/MAX; a single-point value goes in VALUE.
  * Input cells are blue-on-pale; calculated cells are plain; published values
    quoted for comparison are grey italic; check cells are conditionally
    formatted PASS/FAIL.
  * No literal constant is buried inside a formula — constants live on
    'Design Point' and are referenced by defined name.

MIT licensed, per the repository scope map (LICENSE.md).
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName

REV = "1.1"
DOCSET = "lineage v1.2 (docs 00 v1.2 · 10 v1.1 · 11 v1.1 · 12 v1.4 · 20 v1.1 · 21 v1.0 · 22 v1.1 · 30 v1.2 · 31 v1.0 · 40 v1.4 · 50 v1.2)"
REPO = "github.com/phicyclist/slice-cooling"
CONCEPT_DOI = "10.5281/zenodo.21544099"
VERSION_DOI = "10.5281/zenodo.21544100"

# ---------------------------------------------------------------- palette ---
NAVY = "1F3864"; STEEL = "2F5496"; PALE = "DCE6F1"; GREY = "F2F2F2"
INK = "203040"; MUTE = "808080"; OKG = "C6EFCE"; OKT = "006100"
BADF = "FFC7CE"; BADT = "9C0006"; WARN = "FFEB9C"; WARNT = "9C6500"
RULE = Side(style="thin", color="BFBFBF")
BOX = Border(left=RULE, right=RULE, top=RULE, bottom=RULE)


def styles(wb):
    def add(name, **kw):
        st = NamedStyle(name=name)
        st.font = kw.get("font", Font(name="Calibri", size=10, color=INK))
        if "fill" in kw:
            st.fill = PatternFill("solid", fgColor=kw["fill"])
        st.alignment = kw.get("align", Alignment(vertical="top", wrap_text=True))
        st.border = kw.get("border", BOX)
        if "fmt" in kw:
            st.number_format = kw["fmt"]
        wb.add_named_style(st)
        return st

    add("s_hdr", font=Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        fill=STEEL, align=Alignment(vertical="center", wrap_text=True))
    add("s_title", font=Font(name="Calibri", size=16, bold=True, color=NAVY),
        align=Alignment(vertical="center"), border=Border())
    add("s_sub", font=Font(name="Calibri", size=11, bold=True, color=STEEL),
        align=Alignment(vertical="center"), border=Border())
    add("s_txt")
    add("s_num", fmt="0.###", align=Alignment(vertical="top", horizontal="right"))
    add("s_in", font=Font(name="Calibri", size=10, bold=True, color="0033AA"),
        fill=PALE, fmt="0.###", align=Alignment(vertical="top", horizontal="right"))
    add("s_calc", fmt="0.###", align=Alignment(vertical="top", horizontal="right"))
    add("s_ref", font=Font(name="Calibri", size=10, italic=True, color=MUTE),
        fmt="0.###", align=Alignment(vertical="top", horizontal="right"))
    add("s_lbl", font=Font(name="Calibri", size=10, bold=True, color=INK))
    add("s_note", font=Font(name="Calibri", size=9, italic=True, color=MUTE),
        border=Border())
    add("s_sect", font=Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        fill=NAVY, align=Alignment(vertical="center"))


def header(ws, cols, row=1, freeze=True):
    """cols: list of (title, width). Writes a styled header row."""
    for i, (t, w) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=t)
        c.style = "s_hdr"
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def band(ws, row, text, span):
    c = ws.cell(row=row, column=1, value=text)
    c.style = "s_sect"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 16


def titleblock(ws, title, subtitle, span=8):
    ws.cell(row=1, column=1, value=title).style = "s_title"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(row=2, column=1, value=subtitle).style = "s_note"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.row_dimensions[1].height = 22


# =========================================================== REGISTER DATA ===
# (id, track, category, parameter, symbol, value, min, max, unit, grade, gate,
#  basis/derivation, source)
N = None
REGISTER = [
 # ---- design point & ambient ------------------------------------------------
 ("DP-001","Shared","Design point","Ambient dry-bulb (DP-A)","T_amb",32,N,N,"°C","governing","","Sole continuous-duty maximum rating point for every mode","00 §2"),
 ("DP-002","Shared","Design point","Ambient relative humidity (DP-A)","RH_amb",80,N,N,"%","governing","","","00 §2"),
 ("DP-003","Shared","Design point","Ambient humidity ratio","ω_amb",24.2,N,N,"g/kg","governing","","Psat 4.76 kPa → Pv 3.81 → 0.622·3.81/97.5","12 §1"),
 ("DP-004","Shared","Design point","Ambient dew point","T_dp",28.1,N,N,"°C","governing","","Sits at/above every ambient sink — the floor a desiccant breaks","00 §2"),
 ("DP-005","Shared","Design point","Ambient wet-bulb","T_wb",29.5,N,N,"°C","sizing-grade","","","20 §1"),
 ("DP-006","Shared","Design point","Raw-water sink temperature","T_sink",29,N,N,"°C","governing","","Tropical surface/mixed layer; land sinks often cooler = margin","00 §1"),
 ("DP-007","Shared","Design point","Atmospheric pressure","P_atm",101.325,N,N,"kPa","reference","","Magnus basis for all ω conversions","00 §2"),
 ("DP-008","Shared","Design point","Conditioned volume (reference)","V_cab",100,N,N,"m³","platform","","Yacht interior / small dwelling / cabin / shelter","20 §1"),
 ("DP-009","Shared","Design point","Occupants (reference)","n",4,N,N,"person","platform","","","20 §1"),
 ("DP-010","Shared","Design point","Retired secondary design point","—",N,30,75,"°C / %","historical","","Retired; numbers derived from it survive only in archived docs","00 §2"),
 ("DP-011","Shared","Psychrometrics","Saturation humidity ratio at 26 °C","ω_sat",21.3,N,N,"g/kg","reference","","Magnus, 101.325 kPa","00 §2"),
 ("DP-012","Shared","Psychrometrics","Saturation humidity ratio at 27 °C","ω_sat",22.6,N,N,"g/kg","reference","","","00 §2"),
 ("DP-013","Shared","Psychrometrics","Saturation humidity ratio at 28 °C","ω_sat",24.1,N,N,"g/kg","reference","","","00 §2"),
 ("DP-014","Shared","Psychrometrics","Saturation humidity ratio at 29 °C","ω_sat",25.6,N,N,"g/kg","reference","","Working-air exhaust state in the solid balance","00 §2"),
 ("DP-015","Shared","Psychrometrics","Saturation humidity ratio at 30 °C","ω_sat",27.1,N,N,"g/kg","reference","","","00 §2"),
 ("DP-016","Shared","Loads","Occupant latent generation (average)","m_lat",70,50,90,"g/h·person","sizing-grade","","50 sleeping / 90 active; ±30% on load","10 §1; 12 §3"),
 ("DP-017","Shared","Loads","Cabin latent gains, 4 adults","G_lat",280,N,N,"g/h","sizing-grade","","4 × 70 g/h·person","12 §1"),
 ("DP-018","Solid","Loads","Peak sensible load (solar noon)","Q_sens",3.5,N,N,"kW","sizing-grade","","~100 m³ envelope at DP-A","20 §1"),
 ("DP-019","Solid","Loads","Envelope + occupant latent","G_lat",2.0,1.8,2.8,"kg/h","sizing-grade","","Central 2.0; the naive (and wrong) sizing basis — see F1","20 §1"),
 ("DP-020","Shared","Loads","Infiltration, unattended","ACH",0.1,0.05,0.15,"1/h","sizing-grade","","One vs two contactors turns on this band","12 §3"),
 ("DP-021","Liquid","Comfort","Comfort target, liquid track","T/RH",29,N,N,"°C","requirement","","29 °C / 55% RH → ω 13.8 g/kg","10 §1"),
 ("DP-022","Liquid","Comfort","Comfort target humidity ratio","ω_cab",13.8,N,N,"g/kg","requirement","","Psat 4.01 → Pv 2.20 at 29 °C / 55%","12 §1"),
 ("DP-023","Solid","Comfort","Comfort target, solid track","T",25,N,N,"°C","requirement","","~40–55% RH","20 §1"),
 ("DP-024","Solid","Comfort","Cabin steady state, solid track","ω_cab",9.1,N,N,"g/kg","sizing-grade","T2","~40% RH — drier than target; headroom for a warmer setpoint","20 §5"),
 # ---- liquid track: brine chemistry ----------------------------------------
 ("LQ-001","Liquid","Brine","Working concentration, strong","c_hi",40,N,N,"wt%","sizing-grade","A","Water activity aw ≈ 0.45","10 §2"),
 ("LQ-002","Liquid","Brine","Working concentration, dilute","c_lo",35,N,N,"wt%","sizing-grade","A","","10 §2"),
 ("LQ-003","Liquid","Brine","Water activity, 40 wt%","aw",0.45,N,N,"—","PENDING A","A","Published data ±0.05 aw — the largest single unknown","12 §3"),
 ("LQ-004","Liquid","Brine","Water uptake per kg concentrate","Δm",0.143,N,N,"kg/kg","sizing-grade","","c_hi/c_lo − 1 = 0.40/0.35 − 1","10 §3"),
 ("LQ-005","Liquid","Brine","Absorber outlet floor, 40 wt% @ 30 °C brine","ω_out",11.9,10.0,14.2,"g/kg","PENDING A/I","A, I","Temperature-dependent band: 10.0 @27 °C, 14.2 @33 °C","12 §1"),
 ("LQ-006","Liquid","Brine","Hot-regen concentration","c_hi",N,43,44,"wt%","sizing-grade","D","aw 0.33–0.35 at 85–93 °C regeneration","10 §3"),
 ("LQ-007","Liquid","Brine","Absorber outlet floor, 43–44 wt% @ 30 °C","ω_out",9.0,7.5,9.0,"g/kg","PENDING A","A","7.5 g/kg at 27 °C brine","12 §1"),
 ("LQ-008","Liquid","Brine","Cooling leverage on the floor","∂ω/∂T",N,0.5,0.7,"g/kg·K","sizing-grade","A","∂w_eq/∂T at aw 0.45; land cool-water sinks exploit it free","10 §3"),
 ("LQ-009","Liquid","Brine","Crystallization liquidus, 40 wt%","T_liq",N,12,13,"°C","procurement-grade","","CaCl₂ solubility curve; erratum 5 corrected this from ~5 °C","12 §1"),
 ("LQ-010","Liquid","Brine","Crystallization liquidus, 42 wt%","T_liq",N,18,19,"°C","procurement-grade","","","12 §1"),
 ("LQ-011","Liquid","Brine","Crystallization liquidus, 44 wt%","T_liq",22,N,N,"°C","procurement-grade","","","12 §1"),
 ("LQ-012","Liquid","Safety","High-SG cutoff concentration","c_max",42,N,N,"wt%","requirement","","Interlock: unless brine ≥22 °C guaranteed (safety register 6)","11 §4"),
 ("LQ-013","Liquid","Safety","Stored concentrate cap","c_store",43,N,N,"wt%","requirement","","","11 §4"),
 ("LQ-014","Liquid","Brine","Salt specification","—",N,94,97,"%","procurement-grade","","Plain CaCl₂ pellet 94–97% or flake 83–87%; no ferrocyanide/dye/MgCl₂","11 §1"),
 # ---- liquid track: air path and duty ---------------------------------------
 ("LQ-020","Liquid","Air path","Fresh-air floor, 4 occupants","V_fresh",48,N,N,"m³/h","requirement","","12 m³/h·person, CO₂-interlocked; ≈58 kg/h","12 §1"),
 ("LQ-021","Liquid","Air path","Recirculation flow (mixed mode)","V_rec",75,N,N,"m³/h","sizing-grade","I","Decouples removal capacity from the ventilation ration (X6)","10 §2"),
 ("LQ-022","Liquid","Air path","Total absorber flow (mixed mode)","V_tot",123,N,N,"m³/h","sizing-grade","I","≈147 kg/h; = gains 280 g/h ÷ (13.8 − 11.9) g/kg","12 §1"),
 ("LQ-023","Liquid","Air path","ERV latent effectiveness","ε_lat",0.8,N,N,"—","PENDING E","E","Purchasable with membrane area — oversize it","11 §5"),
 ("LQ-024","Liquid","Air path","ERV pre-dried fresh air","ω_ERV",15.9,N,N,"g/kg","sizing-grade","E","ω_amb − ε(ω_amb − ω_cab) at ε 0.8 — see CHK-002","12 §1"),
 ("LQ-025","Liquid","Air path","ERV CO₂ crossover (EATR) limit","EATR",5,N,N,"%","PENDING E","E","Acceptance limit, not a prediction","11 §5"),
 ("LQ-026","Liquid","Duty","Peak removal, 4 adults, mixed mode (bare fresh)","ṁ_w",0.88,N,N,"kg/h","sizing-grade","I","58 kg/h × (24.2−11.9) + 90.6 kg/h × (13.8−11.9)","12 §1"),
 ("LQ-027","Liquid","Duty","Peak removal, ERV'd steady","ṁ_w",0.49,0.49,0.60,"kg/h","sizing-grade","E, I","Quoted 0.49 corresponds to ω_ERV ≈ 17.4 g/kg — see CHK-002","12 §1"),
 ("LQ-028","Liquid","Duty","Regeneration latent duty","Q_lat",0.60,N,N,"kW","sizing-grade","","0.88 kg/h × 2.44 MJ/kg","10 §3"),
 ("LQ-029","Liquid","Duty","Regeneration heat, peak (bare fresh)","Q_regen",0.92,N,N,"kW","sizing-grade","D","latent ÷ COP 0.55–0.75","12 §1"),
 ("LQ-030","Liquid","Duty","Regeneration heat, peak (ERV'd)","Q_regen",0.6,N,N,"kW","sizing-grade","D, E","","12 §1"),
 ("LQ-031","Liquid","Duty","Daily regeneration heat, bare fresh","E_day",N,17,20,"kWh/day","sizing-grade","D","Duty-scheduled at DP-A","12 §1"),
 ("LQ-032","Liquid","Duty","Daily regeneration heat, ERV + DCV","E_day",N,9,11,"kWh/day","sizing-grade","D, E","The budget the solar array is sized to","12 §1"),
 ("LQ-033","Liquid","Duty","Regeneration thermal COP","COP_th",0.59,0.55,0.75,"—","sizing-grade","D","Build-up: latent 0.678 + brine sens. 0.184 + air sens. 0.131 kWh/kg ×1.15","12 §1"),
 ("LQ-034","Liquid","Duty","Regeneration COP with recovery HX","COP_th",0.65,N,N,"—","sizing-grade","D","","12 §1"),
 ("LQ-035","Liquid","Duty","Absorption heat rejected to brine","Q_abs",0.66,N,N,"kW","sizing-grade","","0.88 kg/h × 2.7 MJ/kg — rejection is required, not optional (erratum 2)","12 §1"),
 ("LQ-036","Liquid","Duty","Uncooled contactor self-heating","ΔT",N,5,8,"K","sizing-grade","I","Drives the floor to 12–14+ g/kg","12 §1"),
 ("LQ-037","Liquid","Duty","Air-side heat removal capacity","UA_air",16,N,N,"W/K","sizing-grade","","Per 48 m³/h fresh (≈35 W/K quoted at mixed-mode flow, 10 §3)","12 §1"),
 ("LQ-038","Liquid","Duty","Brine transfer flow, average","V̇_b",0.07,0.04,0.35,"L/min","sizing-grade","","removal ÷ 0.143 ÷ ρ; 0.35 L/min peak","10 §3"),
 ("LQ-039","Liquid","Storage","Overnight reserve, 4 adults","m_res",N,35,40,"kg","sizing-grade","","12 h duty ÷ 0.143 kg/kg, duty-averaged","10 §3"),
 ("LQ-040","Liquid","Storage","Reserve specification band","m_res",N,25,55,"kg","requirement","","Tank sizing spec; vented HDPE 25–60 L, tied down","10 §3; 11 §6"),
 ("LQ-041","Liquid","Water","Still distillate output","V̇_dist",N,8,18,"L/day","sizing-grade","G","Byproduct of regeneration; feeds the M-cycle and tanks","30 §5"),
 # ---- liquid track: hardware ------------------------------------------------
 ("LQ-050","Liquid","Contactor","Cell fill envelope","H×W×D",N,N,N,"mm","sizing-grade","I","600 H × 300 W × 150 D, cross-flow, 300 mm air path","11 §2"),
 ("LQ-051","Liquid","Contactor","Cell rated capacity","ṁ_cell",N,0.4,0.8,"kg/h","PENDING I","I","At 0.6–1 m/s face velocity — the band test I closes","11 §2"),
 ("LQ-052","Liquid","Contactor","Face velocity, design band","v_face",N,0.6,1.0,"m/s","sizing-grade","I","","11 §2"),
 ("LQ-053","Liquid","Contactor","Irrigation rate per cell","V̇_irr",N,7,11,"L/min","sizing-grade","I","150–240 L/min·m²; underwetting collapses K·a nonlinearly","11 §2"),
 ("LQ-054","Liquid","Contactor","Internal recirculation ratio","—",50,N,N,"passes","sizing-grade","I","Per unit concentration change — floor set by sump, not per-pass","11 §2"),
 ("LQ-055","Liquid","Contactor","Mass-transfer coefficient (literature)","K·a",N,1,3,"kg/m³·s","PENDING I","I","Spread sets contactor depth at 1–3 stages","12 §3"),
 ("LQ-056","Liquid","Contactor","NTU for 85% approach","NTU",1.9,N,N,"—","sizing-grade","I","Central estimate two 300 mm stages; conservative three","11 §2"),
 ("LQ-057","Liquid","Contactor","Cabinet envelope","L×W×H",N,N,N,"m","sizing-grade","","0.25 × 0.25 × 1.1 — slim-locker/closet class","11 §2"),
 ("LQ-058","Liquid","Contactor","Air-side ΔP, dry fill","ΔP",N,20,40,"Pa","sizing-grade","I","Turndown state (brine valved off)","11 §2"),
 ("LQ-059","Liquid","Contactor","Air-side ΔP, film (trickle) mode","ΔP",N,5,100,"Pa","sizing-grade","I","vs 1.4–2.8 kPa for bubble columns — the ~90% electrical cut","10 §4"),
 ("LQ-060","Liquid","Contactor","Air-side ΔP, flooded (storm) mode","ΔP",N,100,300,"Pa","sizing-grade","I","Level setpoint, not a gimbal; pool-class tilt tolerance","11 §2"),
 ("LQ-061","Liquid","Contactor","Drift-eliminator carryover","—",N,0.001,0.005,"% of liquid","procurement-grade","B","Chevron PVC; then PP mesh demister. No metal mesh at any grade","11 §3"),
 ("LQ-062","Liquid","Regenerator","Sealed still pool area","A_pool",N,0.3,0.5,"m²","sizing-grade","G","Shallow HDPE/PP/CPVC tray, lid sloped ≥15°","11 §4"),
 ("LQ-063","Liquid","Regenerator","Sealed still specific rate","ṁ\"",N,1,2,"kg/h·m²","PENDING G","G","At design ΔT","11 §4"),
 ("LQ-064","Liquid","Regenerator","Still tray footprint at 0.9 kg/h","A_pool",N,0.5,0.9,"m²","sizing-grade","G","","12 §1"),
 ("LQ-065","Liquid","Regenerator","Pool vapour pressure, 60–93 °C","P_v",N,14,70,"kPa","reference","","vs sink-cooled condenser 4–5 kPa → ΔPv 10–65 kPa (X2)","11 §4"),
 ("LQ-066","Liquid","Regenerator","Condenser vapour pressure","P_v",N,4,5,"kPa","reference","","Sink-cooled at 29 °C","11 §4"),
 ("LQ-067","Liquid","Regenerator","Pool temperature band","T_pool",N,60,93,"°C","sizing-grade","D","Rate lever, not an equilibrium wall (X2)","11 §4"),
 ("LQ-068","Liquid","Regenerator","Still rate multiplier at 93 °C","—",1.45,N,N,"× (85 °C)","sizing-grade","D","CPVC's only unique purchase","11 §4"),
 ("LQ-069","Liquid","Regenerator","Still rate multiplier at 70 °C","—",0.4,N,N,"× (85 °C)","sizing-grade","D","PP / HDPE legal","11 §4"),
 ("LQ-070","Liquid","Regenerator","Still rate multiplier at 60 °C","—",N,0.1,0.17,"× (85 °C)","sizing-grade","D","PVC (non-pressurized tray); holds mothball / 2-person duty","11 §4"),
 ("LQ-071","Liquid","Regenerator","Air-swept exhaust humidity (degraded)","ω_exh",98,N,N,"g/kg","sizing-grade","G","Dew point 53 °C — rains salty condensate below that (erratum 6)","12 §1"),
 ("LQ-072","Liquid","Regenerator","Air-swept condenser recovery","Δω",64,N,N,"g/kg","sizing-grade","G","At 34 °C approach; condensate technical-grade PENDING TDS","12 §1"),
 ("LQ-073","Liquid","Cooling","Raw-water flow, base","V̇_rw",300,N,N,"L/h","sizing-grade","","0.66 kW rejection at 2 K rise","12 §1"),
 ("LQ-074","Liquid","Cooling","Raw-water flow, heat-rich","V̇_rw",N,600,900,"L/h","sizing-grade","","Up to ~2 kW rejection","12 §1"),
 ("LQ-075","Liquid","Cooling","Floor penalty of 3–4 K rise","Δω",1,N,N,"g/kg","sizing-grade","","Halves raw-water flow — the trade is explicit","11 §8"),
 ("LQ-076","Liquid","Electrical","Total electrical, film primary (initial)","P_e",N,25,50,"W","sizing-grade","","Fans 25–40 + recirc 15–30 + raw water 5–15 + peristaltic","12 §1"),
 ("LQ-077","Liquid","Electrical","Total electrical, film primary (scaled)","P_e",N,50,80,"W","sizing-grade","","+ ERV / exhaust fan","12 §1"),
 ("LQ-078","Liquid","Electrical","Raw-water circuit share","P_e",N,30,50,"W","sizing-grade","","Largest single consumer at scale","11 §8"),
 ("LQ-079","Liquid","Electrical","Column annex, drilled rings","P_e",N,100,140,"W","sizing-grade","C","The reason columns are the annex, not the primary","12 §1"),
 ("LQ-080","Liquid","Electrical","Column annex, membrane diffusers","P_e",N,170,380,"W","sizing-grade","C","Avoid — diffuser dynamic wet pressure (erratum 3)","12 §1"),
 ("LQ-081","Liquid","Columns","Superficial air velocity limit","v_s",0.2,N,N,"m/s","sizing-grade","C","→ 13.4 m³/h per 6″ column","11 §9"),
 ("LQ-082","Liquid","Columns","Sparger static head","ΔP",1.37,1.35,1.40,"kPa/10 cm","measured (lit.)","C","At SG 1.4 (erratum 7)","11 §9"),
 ("LQ-083","Liquid","Columns","Membrane diffuser dynamic wet pressure","ΔP",N,1.5,4,"kPa","procurement-grade","C","Rises with fouling; silently doubles blower watts","11 §9"),
 ("LQ-084","Liquid","Columns","Drilled-ring sparger ΔP","ΔP",N,0.3,0.5,"kPa","sizing-grade","C","1 mm holes in CPVC — wins unless approach fraction says otherwise","11 §9"),
 ("LQ-085","Liquid","M-cycle","Berth-cascade supply temperature","T_sup",22.6,N,N,"°C","sizing-grade","H","Working air ⅓ drawn from cabin, exhausted saturated ~27 °C","12 §1"),
 ("LQ-086","Liquid","M-cycle","Berth-cascade sensible output","Q_out",102,N,N,"W/outlet","sizing-grade","H","vs 86 W for the rejected dry-channel bleed (X5)","12 §1"),
 ("LQ-087","Liquid","M-cycle","Berth-cascade water draw","V̇_w",4.1,N,N,"L/day","sizing-grade","H","≤ still distillate — the loop closes","12 §1"),
 ("LQ-088","Liquid","M-cycle","Dew-point effectiveness","ε_dp",N,0.65,0.80,"—","sizing-grade","H","","10 §2"),
 ("LQ-089","Liquid","M-cycle","Supply from 11.9 g/kg feed","T_sup",N,21,23,"°C","sizing-grade","H","Feed dew point 16.7 °C","10 §2"),
 ("LQ-090","Liquid","M-cycle","Supply from 9.0 g/kg feed","T_sup",N,18,21,"°C","sizing-grade","H","Feed dew point 12.5 °C (hot-regen brine)","10 §2"),
 ("LQ-091","Liquid","M-cycle","Supply from 7.5–8 g/kg feed","T_sup",N,15,18,"°C","sizing-grade","H","Feed dew point ~10 °C (LiCl blend)","10 §2"),
 ("LQ-092","Liquid","Thermal bus","Buffer tank volume","V_buf",N,30,60,"L","sizing-grade","","Insulated, 65–90 °C; merges all sources, doubles as DHW","11 §7"),
 ("LQ-093","Liquid","Thermal bus","Thermostatic limit valve setpoint","T_lim",N,85,90,"°C","requirement","","Protects the CPVC/PP hot side","11 §7"),
 ("LQ-094","Liquid","Thermal bus","Solar collector area (ERV'd budget)","A_sol",N,2.5,4.5,"m²","sizing-grade","","9–11 kWh/day ÷ 2.5–3 kWh/m²·day; ~2 m² ETC top-lift with PVT","12 §1"),
 ("LQ-095","Liquid","Thermal bus","Diesel hydronic fuel, full duty","V̇_f",N,1.1,1.4,"L/day","sizing-grade","","÷ 8.5 kWh/L, 5 kW heater class","12 §1"),
]

REGISTER += [
 # ---- volume-threshold ladder (00 §6) ---------------------------------------
 ("VL-001","Liquid","Volume ladder","Once-through cascade M-cycle cooling","V",N,2,5,"m³/outlet","sizing-grade","H","74–112 W; limit = dew point of working air","00 §6"),
 ("VL-002","Liquid","Volume ladder","Unattended, 1 contactor, holding 13 g/kg","V",90,N,N,"m³","sizing-grade","I","×2 per contactor; limit = contactor capacity vs 0.1 ACH","00 §6"),
 ("VL-003","Liquid","Volume ladder","Unattended, 1 contactor, mold-safe","V",N,160,215,"m³","sizing-grade","I","60–65% RH basis, ~30–50 W + trickle heat","00 §6"),
 ("VL-004","Liquid","Volume ladder","Once-through occupied (pressure integrity)","V",N,110,240,"m³","sizing-grade","I","Supply ≥ 2–3× infiltration","00 §6"),
 ("VL-005","Liquid","Volume ladder","Mixed-mode occupied (baseline)","V",N,300,500,"m³","sizing-grade","I","With a 3-cell bank; limit = bank capacity + heat budget","00 §6"),
 ("VL-006","Liquid","Volume ladder","Whole-cabin AC (recirc + hot-regen brine)","V",100,N,N,"m³","deferred","","~10.6 kg/h, ~11 kW heat — converges with the solid track (X1)","00 §6"),
 ("VL-007","Liquid","Volume ladder","Whole-cabin once-through absorber duty","ṁ_w",N,5.5,11,"kg/h","sizing-grade","","Drying 370–750 m³/h of ambient supply — why X1 holds","12 §1"),
 ("VL-008","Liquid","Performance","4 adults, once-through, base brine","RH",N,59,66,"%","sizing-grade","I","FAILS comfort — mode retired for full occupancy (erratum 8)","10 §5"),
 ("VL-009","Liquid","Performance","4 adults, mixed-mode, cooled base brine","RH",55,N,N,"%","sizing-grade","I","The occupied baseline (X6), at 29 °C","10 §5"),
 ("VL-010","Liquid","Performance","2 adults, once-through, cooled brine","RH",N,54,58,"%","sizing-grade","I","Stands on the base system","10 §5"),
 # ---- solid track -----------------------------------------------------------
 ("SD-001","Solid","Duty","Total peak sorbent duty (self-consistent)","ṁ_w",N,9,11,"kg/h","sizing-grade","T2","F1: working-air moisture dominates; sharpened from 7–8 kg/h","20 §5"),
 ("SD-002","Solid","Duty","Naive envelope-only duty (cautionary row)","ṁ_w",N,2,3,"kg/h","superseded","","Under-sizes the desiccant 3–5× — retained as the error record","22 §4"),
 ("SD-003","Solid","Duty","Continuous regeneration heat","Q_regen",N,7.5,9,"kW","sizing-grade","M2, T2","~0.85 kWh/L × duty","20 §5"),
 ("SD-004","Solid","Duty","Parasitic electrical","P_e",N,0.6,1.0,"kW","PENDING M1","M1","Fans 300–660 W supply + 60–145 W working; pumps 100–250 W","20 §5"),
 ("SD-005","Solid","Duty","Supply fan pressure","ΔP",N,250,550,"Pa","PENDING M1","M1","Fan efficiency η 0.32 assumed","20 §5"),
 ("SD-006","Solid","Duty","Working fan pressure","ΔP",N,150,350,"Pa","PENDING M1","M1","","20 §5"),
 ("SD-007","Solid","Air path","Supply airflow","ṁ_a",N,0.43,0.51,"kg/s","sizing-grade","T2","≈1,350–1,590 m³/h; topology-dependent (dry-draw vs cabin-draw)","20 §5"),
 ("SD-008","Solid","Air path","Working airflow","ṁ_w,air",N,0.13,0.16,"kg/s","sizing-grade","T2","400–500 m³/h; Q_wet ÷ Δh","20 §5"),
 ("SD-009","Solid","Air path","Supply temperature","T_sup",N,16.8,18.1,"°C","sizing-grade","T2","Set by working-air dew point at ε_dp 0.7","20 §5"),
 ("SD-010","Solid","Air path","Post-desiccant humidity ratio","ω_sup",8,N,N,"g/kg","sizing-grade","M1","AlFu back-end; state 1–3 of the chain","20 §2"),
 ("SD-011","Solid","Air path","Post-desiccant air temperature","T",N,50,55,"°C","sizing-grade","M1","Adsorption heat partly removed by sink fluid","20 §2"),
 ("SD-012","Solid","Air path","Post-intercool temperature","T",N,30,32,"°C","sizing-grade","","Sensible only, to raw water","20 §2"),
 ("SD-013","Solid","Air path","Working exhaust state (recycled)","ω",25.6,N,N,"g/kg","sizing-grade","","Saturated at 29 °C; never vented in X8 mode","20 §2"),
 ("SD-014","Solid","Air path","M-cycle dew-point effectiveness","ε_dp",0.7,N,N,"—","sizing-grade","H","Toward ~10.8 °C dew point on dried air","20 §5"),
 ("SD-015","Solid","Air path","Post-ERV fresh state","T/ω",31,12,N,"°C / g/kg","sizing-grade","E","Pre-dried from 24.2 g/kg at ε 0.8","20 §2"),
 ("SD-016","Solid","Cycle","Half-cycle time","t_half",10,N,N,"min","sizing-grade","M2","Throughput is governed by cycle time and coated area, not mass","22 §1"),
 ("SD-017","Solid","Energy","Latent floor of desorption","e_lat",0.67,N,N,"kWh/L","reference","","Heat of vaporization — material-independent, 70–85% of input","20 §3"),
 ("SD-018","Solid","Energy","Binding excess (sorbent-dependent)","e_bind",N,0.05,0.30,"kWh/L","sizing-grade","M3","MOF low","20 §3"),
 ("SD-019","Solid","Energy","Sensible bed cycling","e_sens",N,0.03,0.10,"kWh/L","PENDING M3","M3","May be optimistic for the plumbing, not just the bed (doc 20 §4)","20 §3"),
 ("SD-020","Solid","Energy","Total specific regeneration energy","e_tot",N,0.8,1.0,"kWh/L","PENDING M3","M3","The M3 acceptance figure","20 §3"),
 ("SD-021","Solid","Energy","Recoverable fraction of input","—",N,10,15,"%","sizing-grade","","Condenser latent is downhill and DHW-demand-capped (1–3 kWh/day)","20 §3"),
 ("SD-022","Solid","Water","Regeneration condensate","V̇_c",N,150,220,"L/day","sizing-grade","T2","Covers the M-cycle wet-channel feed","20 §5"),
 ("SD-023","Solid","Water","Potable-grade surplus","V̇_s",N,50,70,"L/day","sizing-grade","T2","The gains + ventilation terms; water-neutral M-cycle (X8)","20 §5"),
 ("SD-024","Solid","Regeneration","Design regeneration temperature","T_reg",N,60,65,"°C","sizing-grade","M2","Purge-humidity-dependent — the most misquoted figure in the literature","20 §6"),
 ("SD-025","Solid","Regeneration","Representative purge humidity","ω_purge",25,24,25,"g/kg","sizing-grade","M2","M2 is run against ~24 g/kg logged purge, never dry lab air","20 §6"),
 ("SD-026","Solid","Regeneration","Purge RH at bed face, 45 °C","RH_face",38,N,N,"%","sizing-grade","M2","Above AlFu's step — zero driving force (F2)","20 §6"),
 ("SD-027","Solid","Regeneration","Purge RH at bed face, 50 °C","RH_face",29,N,N,"%","sizing-grade","M2","On the step — marginal","20 §6"),
 ("SD-028","Solid","Regeneration","Purge RH at bed face, 60 °C","RH_face",17,N,N,"%","sizing-grade","M2","Below the step — works","20 §6"),
 ("SD-029","Solid","Regeneration","Purge RH at bed face, 65 °C","RH_face",13,N,N,"%","sizing-grade","M2","Below the step — comfortable","20 §6"),
 ("SD-030","Solid","Sorbent","AlFu isotherm step position","RH_step",N,25,30,"%","procurement-grade","T4","Type V; the decisive selection property","21 §2"),
 ("SD-031","Solid","Sorbent","Working Δq, dry-purge full swing","Δq",N,0.2,0.3,"g/g","procurement-grade","T4","F3: this is NOT the sizing number — see SD-032","21 §2"),
 ("SD-032","Solid","Sorbent","Effective Δq under humid purge","Δq_eff",N,0.15,0.2,"g/g","PENDING M2","M2","F3: residual 0.05–0.10 g/g at 60–65 °C → +25–50% inventory","21 §6"),
 ("SD-033","Solid","Sorbent","Residual loading at 60–65 °C","q_res",N,0.05,0.10,"g/g","PENDING M2","M2","Against ~25 g/kg purge","21 §6"),
 ("SD-034","Solid","Sorbent","BET surface area, aqueous route","S_BET",1135,1000,1135,"m²/g","procurement-grade","T4","QC gate 4: ≥1,000; target 1,100–1,135","21 §6"),
 ("SD-035","Solid","Sorbent","BET surface area, grinding route","S_BET",N,750,1100,"m²/g","PENDING T4","T4","Gate ≥750–800; expected 790–1,100","21 §4"),
 ("SD-036","Solid","Sorbent","Yield gate (activated dry mass)","Y",90,N,N,"% theor.","requirement","T4","~158 g/mol per Al","21 §6"),
 ("SD-037","Solid","Sorbent","Demonstrated cycle life, AlFu","N_cyc",4500,N,N,"cycles","measured (lit.)","M4","Unchanged; sibling CAU-10-H to 10,000. Failure mode is mechanical","21 §2"),
 ("SD-038","Solid","Coating","Coating thickness","t",N,0.1,0.5,"mm","sizing-grade","M1","Sprayed (preferred over dip) onto prepared aluminium","22 §2"),
 ("SD-039","Solid","Coating","Sorbent loading, planform","m\"",0.18,N,N,"kg/m²","sizing-grade","M1","","22 §2"),
 ("SD-040","Solid","Coating","PVA hydrolysis grade","—",N,98,99,"%","requirement","M4","Fully hydrolyzed only — partial grades redissolve in humid service","22 §2"),
 ("SD-041","Solid","Coating","Binder fraction","w_b",10,N,N,"wt%","sizing-grade","M4","Nominal; locked by test","22 §2"),
 ("SD-042","Solid","Coating","Anneal / activation temperature","T_act",N,120,150,"°C","requirement","","6–12 h; doubles as the mandatory PVA anneal — a re-coat never skips it","22 §2"),
 ("SD-043","Solid","Sizing","Sorbent inventory, envelope-only basis","m_s",N,3,4,"kg","superseded","T2","Cautionary row; real charge is 3–5× this, then +25–50% for F3","22 §4"),
 ("SD-044","Solid","Sizing","Coated HX area, envelope-only basis","A_c",N,10,13,"m²","superseded","T2","Scales with the corrected duty","22 §4"),
 ("SD-045","Solid","Sizing","Inventory multiplier, self-consistent duty","—",N,3,5,"×","sizing-grade","T2","Over the envelope-only row","22 §4"),
 ("SD-046","Solid","Sizing","Inventory multiplier, F3 compounded","—",N,1.25,1.50,"×","PENDING M2","M2","Applied on top of SD-045","22 §4"),
 ("SD-047","Solid","Hardware","DCHX volumetric duty","—",90,N,N,"kW/m³","reference","","vs ~59 for packed granular beds; 3–5× heat-transfer rate","22 §1"),
 ("SD-048","Solid","Synthesis","Al₂(SO₄)₃·18H₂O, bench batch","m",166.6,N,N,"g","procurement-grade","T1","MW 666.43; 0.50 mol Al → ~75 g MOF (validation 41.7 g / full 5.5 kg)","21 §3"),
 ("SD-049","Solid","Synthesis","Fumaric acid, bench batch","m",58.0,N,N,"g","procurement-grade","T1","MW 116.07; 0.50 mol (validation 14.5 g / full 1.9 kg)","21 §3"),
 ("SD-050","Solid","Synthesis","NaOH, bench batch","m",40.0,N,N,"g","procurement-grade","T1","MW 40.00; 1.00 mol → 1:1:2 mole check (validation 10 g / full 1.3 kg)","21 §3"),
 ("SD-051","Solid","Synthesis","DI water, bench batch","V",750,N,N,"mL","procurement-grade","T1","~190 mL for the validation batch","21 §3"),
 ("SD-052","Solid","Synthesis","Precipitation temperature","T",60,N,N,"°C","procurement-grade","T1","Hold 30–60 min; wash 3–4× DI; dry ~100 °C","21 §3"),
 ("SD-053","Solid","Synthesis","LAG liquid ratio","η",N,0.1,1,"µL/mg","developmental","T1","Route B; 1 Al : 1 fumarate, no base","21 §4"),
 ("SD-054","Solid","Synthesis","LAG aging temperature","T_age",N,110,120,"°C","developmental","T1","The aging step does much of the crystal growth","21 §4"),
 ("SD-055","Solid","Synthesis","Ball-to-powder ratio (mill path)","BPR",N,10,30,"×","developmental","T1","Mill path only; mill-zero is mortar-and-pestle","21 §4"),
 ("SD-056","Solid","Synthesis","Full-charge synthesis scale","m",2.5,N,N,"kg","indicative","T2","Mill returns at scale-up — hand-grinding does not scale","21 §3"),
 ("SD-057","Solid","Bench rig","Coupon loop flow","V̇",N,0.5,1.5,"L/min","procurement-grade","M3","Hot-water-rated DC pump; all-Al + SS wetted path, distilled water","22 §5"),
 ("SD-058","Solid","Bench rig","Hot reservoir range","T",N,50,90,"°C","procurement-grade","M2","Sous-vide circulator as the precision hot source","22 §5"),
 ("SD-059","Solid","Bench rig","Cold reservoir range","T",N,10,25,"°C","procurement-grade","M1","Ambient / ice-bathed","22 §5"),
 ("SD-060","Solid","Bench rig","Capacitive RH sensor validity ceiling","T",60,N,N,"°C","procurement-grade","","Spec'd 20–80% RH; drifts ~+3% RH in sustained saturation → use wet/dry bulb","22 §7"),
 ("SD-061","Solid","Bench rig","Durability cycle count","N_cyc",N,100,1000,"cycles","requirement","M4","Capacity fade + coating mass loss/delamination","22 §6"),
]

REGISTER += [
 # ---- CO2 / ventilation stack (00 §5, spec P17) -----------------------------
 ("CO2-001","Shared","Spec P17","Cabin CO₂ limit, all modes","c_CO2",1000,N,N,"ppm","requirement","J","Safety-critical; includes sealed operation","00 §5"),
 ("CO2-002","Shared","Spec P17","Alarm + forced boost threshold","c_CO2",2000,N,N,"ppm","requirement","","Per-room max sensing governs the interlock","00 §5"),
 ("CO2-003","Shared","Ventilation","CO₂ generation, 4 occupants awake","V̇_CO2",0.072,N,N,"m³/h","reference","","Crew total, not per person (0.047 asleep); ≈3.4 kg/day","00 §5"),
 ("CO2-004","Shared","Ventilation","CO₂ mass generation, 4 occupants","ṁ_CO2",3.4,N,N,"kg/day","reference","","","00 §5"),
 ("CO2-005","Shared","Ventilation","Fresh-air floor","V̇_f",48,N,N,"m³/h","requirement","","12 m³/h·person; mechanical minimum stop on the damper","00 §5"),
 ("CO2-006","Shared","Ventilation","Steady CO₂ at 48 m³/h (awake)","c_CO2",1920,N,N,"ppm","sizing-grade","","ABOVE the <1,000 spec — closed by the layer stack, not by ventilation","00 §5"),
 ("CO2-007","Shared","Ventilation","Steady CO₂ at 48 m³/h (asleep)","c_CO2",1400,N,N,"ppm","sizing-grade","","","12 §1"),
 ("CO2-008","Shared","Ventilation","Steady CO₂ at 36 m³/h","c_CO2",2420,N,N,"ppm","sizing-grade","","Dose–response ladder — why the interlock exists","00 §5"),
 ("CO2-009","Shared","Ventilation","Steady CO₂ at 24 m³/h","c_CO2",3420,N,N,"ppm","sizing-grade","","","00 §5"),
 ("CO2-010","Shared","Ventilation","Steady CO₂ at 12 m³/h","c_CO2",6420,N,N,"ppm","sizing-grade","","","00 §5"),
 ("CO2-011","Shared","Ventilation","Sealed envelope (infiltration only)","c_CO2",7600,N,N,"ppm","sizing-grade","","","00 §5"),
 ("CO2-012","Shared","Ventilation","Closed-room penalty over main space","Δc",N,2000,2900,"ppm","sizing-grade","","Drives displacement ducting + per-room NDIR sensing","00 §5"),
 ("CO2-013","Shared","Ventilation","Marginal heat cost of fresh air (ε 0.8)","e",0.06,N,N,"kWh/day per m³/h","sizing-grade","E","(1−ε)×10.4 g/kg — makes generous ventilation cheap (X7)","00 §5"),
 ("CO2-014","Shared","Ventilation","Heat saved per 12 m³/h cut (bare fresh)","E",3.8,N,N,"kWh/day","sizing-grade","","The standing economic temptation the interlock forecloses","00 §5"),
 ("CO2-015","Shared","CO₂ battery","Scrubbing duty at the ventilation floor","ṁ_CO2",1.6,N,N,"kg/day","PENDING J","J","Two-bed solid-amine TSA in the recirculation branch","00 §5"),
 ("CO2-016","Shared","CO₂ battery","Bed mass, amine (each of two)","m_b",3,N,N,"kg","PENDING J","J","~90-min half-cycles","00 §5"),
 ("CO2-017","Shared","CO₂ battery","Recirculation flow through the bed","V̇",91,N,N,"m³/h","PENDING J","J","At 50% single-pass capture; +50–150 Pa","00 §5"),
 ("CO2-018","Shared","CO₂ battery","Amine regeneration temperature","T_reg",N,85,95,"°C","PENDING J","J","Off the same heat bus; solar-window schedulable","00 §5"),
 ("CO2-019","Shared","CO₂ battery","Specific regeneration energy","e",N,1.0,1.3,"kWh/kg","PENDING J","J","Includes water co-adsorption + sensible","12 §1"),
 ("CO2-020","Shared","CO₂ battery","Daily regeneration heat","E",N,1.6,2.5,"kWh/day","PENDING J","J","","12 §1"),
 ("CO2-021","Shared","CO₂ battery","Overnight stored obligation","m",0.3,N,N,"kg","sizing-grade","J","Loaded beds hold the night → regeneration is schedulable","00 §5"),
 ("CO2-022","Shared","CO₂ battery","Optimal bed placement humidity","RH",N,45,55,"%","procurement-grade","J","Post-dehumidification; near-optimal for amine chemistry","00 §5"),
 ("CO2-023","Shared","CO₂ battery","Potash (K₂CO₃/carbon) bed mass","m_b",N,4,5,"kg","PENDING J-K","J-K","At ~0.5–0.65 mmol/g working capacity","00 §5"),
 ("CO2-024","Shared","CO₂ battery","Potash regeneration temperature","T_reg",N,130,150,"°C","PENDING J-K","J-K","Requires a ≥~130 °C tap; equilibrium-dead below ~120 °C (X11)","00 §5"),
 ("CO2-025","Shared","CO₂ battery","Potash working capacity","q",N,0.5,0.65,"mmol/g","PENDING J-K","J-K","Alumina and MgO supports prohibited — double-salt deactivation","00 §5"),
 ("CO2-026","Shared","Fallback","Oversized ERV + DCV flow","V̇_f",N,120,190,"m³/h","sizing-grade","E","→ 800–1,000 ppm, open conditions only, no sealed guarantee","00 §5"),
 ("CO2-027","Shared","X9 galley","Burner CO₂ emission, 2 kW","V̇_CO2",0.26,N,N,"m³/h","reference","","3.6× the whole crew — the reason for no combustion in the envelope","00 §5"),
 ("CO2-028","Shared","X9 galley","Burner combustion water","ṁ_w",0.25,N,N,"kg/h","reference","","Latent load added on top of the CO₂","00 §5"),
 ("CO2-029","Shared","X9 galley","Induction galley electrical","E_e",2,N,N,"kWh/day","sizing-grade","","Deletes both burner terms","00 §5"),
 ("CO2-030","Shared","Rejected","Raw-water CO₂ absorption flow","V̇",6,N,N,"m³/h","rejected","","Fully equilibrated, slow kinetics, re-humidifies the air","12 §5"),
 ("CO2-031","Shared","Rejected","Membrane driving pressure","ΔP",0.1,N,N,"kPa","rejected","","Cabin-ppm partial pressure — no membrane process closes","12 §5"),
 ("CO2-032","Shared","Rejected","Algae/plant light demand","P",150,N,N,"kWh/day","rejected","","For 3.4 kg/day","12 §5"),
 ("CO2-033","Shared","Rejected","Soda lime / LiOH consumable rate","ṁ",15,N,N,"kg/day","emergency only","","Emergency consumable, never an architecture","12 §5"),
 ("CO2-034","Shared","Rejected","Electrochemical capture (watch item)","e",1,N,N,"kWh_e/kg","R&D watch","","In principle; not procurable at scale","12 §5"),
 ("CO2-035","Shared","Rejected","CALF-20 capacity at cabin ppm","q",N,0.1,0.17,"mmol/g","rejected","","Selectivity ceiling ~40–47% RH straddles the bed placement","12 §5"),
 ("CO2-036","Shared","Rejected","SIFSIX-class capacity / water co-adsorption","q",1.25,7.5,10,"mmol/g","rejected","","ppm-capable at ~1.2–1.3, but 7.5–10 mmol/g water rides every swing","12 §5"),
 ("CO2-037","Shared","Rejected","Mg-MOF-74 humid capacity retention","—",16,N,N,"%","rejected","","Water poisons and hydrolyzes the open metal sites","12 §5"),
 ("CO2-038","Shared","Rejected","Moisture-swing AER displaced heat","E",N,5,15,"kWh/day","rejected","","vs 1.6–2.5 for TSA — an X8 rule-1 violation at DP-A","12 §5"),
 # ---- platform integration (doc 30) -----------------------------------------
 ("IN-002","Platform","Heat cascade","Two-stage HDH humidifier grade","T",N,75,85,"°C","sizing-grade","","Below the potash rung, above the amine rung","30 §2"),
 ("IN-003","Platform","Heat cascade","HDH gained output ratio","GOR",N,2.2,2.8,"—","sizing-grade","","Warm sink pinches the cold end to ~34 °C — lab GOR-4+ does not transfer","30 §3"),
 ("IN-004","Platform","Heat cascade","HDH yield per unit heat","V̇",90,N,N,"L/day·kW","sizing-grade","","Raw-water feed sites only; never pointed at cabin air","30 §3"),
 ("IN-005","Platform","Heat cascade","HDH stranded humidity at pinch","ω",35,N,N,"g/kg","sizing-grade","","Desiccant polish stage recaptures +20–25% when heat is free","30 §3"),
 ("IN-006","Platform","Heat cascade","Resistive-HDH energy penalty vs RO","—",27,N,N,"×","reference","","Clipped-solar dump load and last resort only","30 §2"),
 ("IN-007","Platform","Heat cascade","Marine RO specific energy","e",N,4,20,"Wh/L","reference","","The honest benchmark for deliberately manufactured water","20 §7"),
 ("IN-008","Platform","Sources","PVT thermal grade","T",N,45,50,"°C","procurement-grade","","Cannot reach HDH grade; F2-dead for direct solid-bed regen","30 §2"),
 ("IN-009","Platform","Sources","PVT panel mass","m",36,N,N,"kg","procurement-grade","","Beyond two panels the marginal thermal output has no user","30 §4"),
 ("IN-010","Platform","Sources","ETC collector band","T",N,60,93,"°C","procurement-grade","","The liquid track's full grade band (X2) — the solar comfort island","30 §2"),
 ("IN-011","Platform","Water ladder","Solid-track condensate surplus","V̇",N,50,70,"L/day","sizing-grade","T2","Independent of HDH hardware — a real redundancy path","30 §5"),
 ("IN-012","Platform","Water ladder","Liquid-track still distillate","V̇",N,8,18,"L/day","sizing-grade","G","The solar-independent path; also the M-cycle feed","30 §5"),
 ("IN-013","Platform","Comfort","Conductive cooling pad duty","Q",N,70,90,"W","sizing-grade","","Sleeping; skin 33–35 °C vs 29 °C sink through a Ti/CuNi plate HX","30 §3"),
 ("IN-014","Platform","Raw water","Marine intake depth","d",N,5,10,"m","procurement-grade","","For stability and cleanliness, NOT temperature (mixed layer isothermal to 20–40 m)","30 §3"),
 ("IN-015","Platform","Marine","Design heel / roll","θ",N,15,20,"°","requirement","I, H","Flooded-mode fallbacks, staged headers, anti-slosh rules","00 §1"),
 ("IN-016","Platform","Marine","Passive fluid-path heel rating","θ",30,N,N,"°","requirement","","Thermosyphons, wicked heat pipes, trapped siphons, check-valved backups","30 §7"),
 # The cascade-tap interface — what the comfort systems actually require of any
 # upstream source. The grades themselves are carried by the component rows named
 # in each note; these three record them as the PLATFORM interface requirement.
 ("IN-017","Platform","Interface","Cascade tap — liquid-track sealed still","T",N,60,93,"°C","requirement","D","Interface requirement on any source; grade from LQ-067 / IN-010 (X2)","30, function statement"),
 ("IN-018","Platform","Interface","Cascade tap — solid-track regeneration","T",N,60,65,"°C","requirement","M2","Interface requirement on any source; grade from SD-024 (F2)","30, function statement"),
 ("IN-019","Platform","Interface","Cascade tap — potash CO₂ bed where primary","T",130,N,N,"°C","requirement","J-K","Interface requirement, platform-conditional; grade from CO2-024 (X11). Stated as a floor (≥ ~130 °C)","30, function statement"),
 # ---- upgrade paths (doc 31) ------------------------------------------------
 ("UP-001","Upgrade","X12 AHT","Absorption heat transformer COP","COP",N,0.45,0.48,"—","estimate-grade","L","Upgraded heat ÷ driving heat, before losses","31 §2.4"),
 ("UP-002","Upgrade","X12 AHT","Evaporator temperature","T_ev",N,60,65,"°C","estimate-grade","L","The waste tail itself, boiling distillate at ~20–25 kPa","31 §2.2"),
 ("UP-003","Upgrade","X12 AHT","Absorber (delivered) temperature","T_abs",N,85,90,"°C","estimate-grade","L","Amine-bed regen grade + hot-regen still grade","31 §2.2"),
 ("UP-004","Upgrade","X12 AHT","Gross temperature lift","ΔT",N,20,25,"K","estimate-grade","L","","31 §2.3"),
 ("UP-005","Upgrade","X12 AHT","Lift ceiling, 44 wt% @ T_ev 65 °C","T_abs",91,N,N,"°C","estimate-grade","A, L","aw(x)·Psat(T_abs) < Psat(T_ev); 85–86 °C at T_ev 60 °C","31 §2.3"),
 ("UP-006","Upgrade","X12 AHT","Lift ceiling, 42 wt% @ T_ev 65 °C","T_abs",86.5,86,87,"°C","estimate-grade","A, L","81 °C at T_ev 60 °C","31 §2.3"),
 ("UP-007","Upgrade","X12 AHT","Tail heat per unit upgraded heat","—",2.2,N,N,"kW/kW","estimate-grade","L","~5 kWh/day tail delivers the CO₂ battery's 2–2.5 kWh/day at grade","31 §2.4"),
 ("UP-008","Upgrade","X12 AHT","Distillate lift pump head","ΔP",16,N,N,"kPa","estimate-grade","","Watt-scale peristaltic — unlike LiBr machines' solution pump","31 §2.2"),
 ("UP-009","Upgrade","X12 AHT","Absorption latent release","h",2.6,N,N,"MJ/kg","reference","","Released at the brine's temperature — the whole X12 finding","31 §2.1"),
 ("UP-010","Upgrade","VC heat pump","Electrical draw","P_e",N,350,500,"W","estimate-grade","I","5× step over the 25–80 W baseline — boost mode only","31 §3"),
 ("UP-011","Upgrade","VC heat pump","Evaporator duty","Q_ev",N,0.8,1.0,"kW","estimate-grade","I","0.66 kW absorption heat + parasitic ingress through the contactor","31 §3"),
 ("UP-012","Upgrade","VC heat pump","Condenser output","Q_c",N,1.1,1.5,"kW","estimate-grade","","Covers the 0.92 kW regen peak with surplus into the buffer","31 §3"),
 ("UP-013","Upgrade","VC heat pump","Real heating COP at ~55 K lift","COP_h",N,2.5,3.0,"—","estimate-grade","","Small hermetic, inverter-driven","31 §3"),
 ("UP-014","Upgrade","VC heat pump","Net floor gain at 20 °C brine","Δω",N,3.5,4.5,"g/kg","estimate-grade","A, I","Most of the LiCl/hot-regen benefit for ~400 W and no new chemistry","31 §3"),
 ("UP-015","Upgrade","VC heat pump","Glycol-loop approach penalty","Δω",N,1.5,2.0,"g/kg","estimate-grade","","Cost of the 2–3 K two-worlds isolation approach","31 §3"),
 ("UP-016","Upgrade","VC heat pump","R290 condensing pressure at 70 °C","P",2.6,N,N,"MPa","reference","","Procurement winner: 0.5–1 kW HPWH rotaries; T_crit 96.7 °C","31 §3"),
 ("UP-017","Upgrade","VC heat pump","R600a condensing pressure at 70 °C","P",1.05,N,N,"MPa","reference","","Best thermodynamic fit; charge 150–300 g, A3 flammable","31 §3"),
 ("UP-018","Upgrade","Rejected","Prime-mover heat pump electrical draw","P_e",2,N,N,"kW","rejected","","At ~35 K sink-to-regen lift, COP_h 3.5–4 — loses to a VC-AC outright","31 §1"),
 ("UP-019","Upgrade","Still MVR","Electrical draw at 0.88 kg/h","P_e",N,110,140,"W","estimate-grade","","PV-scale: makes the moisture battery rechargeable through a heat outage","31 §4"),
 ("UP-020","Upgrade","Still MVR","Specific compression work","w",N,120,160,"Wh/kg","estimate-grade","","vs ~1,800 Wh/kg thermal","31 §4"),
 ("UP-021","Upgrade","Still MVR","Headspace pressure, 70 °C / 42 wt%","P",12.5,N,N,"kPa","estimate-grade","","aw ≈ 0.40 — nearly half the pressure ratio is the activity depression","31 §4"),
 ("UP-022","Upgrade","Still MVR","In-pool condensing pressure","P",40,N,N,"kPa","estimate-grade","","Sat. 76–78 °C for a useful 6–8 K ΔT → pressure ratio ~3","31 §4"),
 ("UP-023","Upgrade","Still MVR","Vapour volumetric flow","V̇",N,11,13,"m³/h","estimate-grade","","At 12.5 kPa; PR ~3 excludes regenerative blowers (PR ~1.2 ceiling)","31 §4"),
 ("UP-024","Upgrade","Still MVR","Compression superheat","ΔT",N,80,90,"K","estimate-grade","","Condensate-spray desuperheater ahead of the pool coil","31 §4"),
 ("UP-025","Upgrade","AlFu chiller","Chilled-water cooling duty","Q_c",N,0.9,1.0,"kW","estimate-grade","M-series","Holding brine at ~20 °C","31 §5"),
 ("UP-026","Upgrade","AlFu chiller","Thermal COP","COP_th",N,0.4,0.5,"—","estimate-grade","M-series","→ 2–2.5 kW of 60–90 °C heat; +2–3 m² collector if solar-fed","31 §5"),
 ("UP-027","Upgrade","AlFu chiller","Evaporator absolute pressure","P",N,1.7,2.0,"kPa","estimate-grade","M-series","Leak-tightness is the make-or-break property","31 §5"),
 ("UP-028","Upgrade","Crystallizer","Dihydrate CaCl₂ content","w",75.5,N,N,"wt%","reference","A3","CaCl₂·2H₂O","31 §6"),
 ("UP-029","Upgrade","Crystallizer","Water uptake per kg dihydrate","Δm",N,1.1,1.2,"kg/kg","estimate-grade","A3","vs 0.143 kg/kg for the 40→35 wt% liquid swing","31 §6"),
 ("UP-030","Upgrade","Crystallizer","Reserve mass reduction","—",N,3,5,"×","estimate-grade","A3","Occupied-night charge 35–40 kg → 10–15 kg","31 §6"),
 ("UP-031","Upgrade","Crystallizer","Supercooling without seeding","ΔT",N,10,20,"K","reference","A3","Crystals never travel — the pot is isolated by a valve pair","31 §6"),
 ("UP-032","Upgrade","PCM","CaCl₂·6H₂O melting point","T_m",N,29,30,"°C","reference","","Sits essentially at the sink temperature at DP-A — logged, low priority","31 §6.2"),
]


# ============================================================ TEST PROGRAM ===
# (id, track, title, cost_lo, cost_hi, duration, in_baseline, decides, gates)
TESTS = [
 ("A","Liquid","Jar equilibrium — real aw table",30,30,"days",1,
  "Sealed jars + RH probe over CaCl₂ 35/38/40/42/44 wt% (+LiCl arm) at ~27 and ~33 °C",
  "Go/no-go: 40 wt% @ 33 °C ≤ 55% ERH. Feeds every floor number and the X12 lift inequality"),
 ("A2","Liquid","Aerated hot coupon (formate variant)",20,20,"weeks",0,
  "Whether potassium-formate brine survives aerated 80 °C service",
  "Materials relaxation for that variant only"),
 ("A3","Upgrade","Crystallizer jar extension",10,10,"days",0,
  "Seeded vs unseeded 44 wt% parked at ~25 °C: supercooling, phase, caking over 50–100 cycles",
  "Gates the static crystallizer pot (doc 31 §6); mass-limited platforms only"),
 ("B","Liquid","Aerosol drift — steel coupon",5,5,"1 week",1,
  "Bare mild-steel coupon downstream of eliminator + demister, max face velocity, trickle and flooded",
  "Any rust = redesign before any cabin connection. Also arbitrates X3 (safety register item 3)"),
 ("C","Liquid","Bubble column / sparger shootout",150,150,"weeks",0,
  "Sparger dynamic wet pressure comparison",
  "Only if the column annex is ordered"),
 ("D","Liquid","Regeneration COP",50,50,"days",1,
  "Still tray + air-swept at 60/70/85 °C incl. the PVC-tray 60 °C point and condenser approach",
  "Confirms the 0.55–0.75 COP band and the degraded ladder"),
 ("E","Liquid","ERV core",80,150,"weeks",1,
  "Real ε_lat; salt-aerosol fouling trend (U-tube ΔP); condensation at DP-A inlet; CO₂ crossover",
  "Gates the 9–11 kWh/day heat budget (X7); EATR <5%"),
 ("F","Liquid","Endurance (passive)",0,0,"months",1,
  "Locker/closet-dryer duty: salt creep, fouling, crystallization events, ΔP trend",
  "Runs behind everything else at zero cost"),
 ("G","Liquid","Sealed-still rate",40,40,"days",1,
  "kg/h·m² vs pool temperature; condensate TDS (entrainment ≈ 0); TDS on the air-swept condenser stream",
  "Gates distillate quality claims and the X8 degraded-mode condensate"),
 ("H","Both","M-cycle wetting / heel",60,60,"days",1,
  "Supply temp vs feed humidity; 5/10/15° tilt; closed-loop feed point (X8 recycled dried exhaust)",
  "Shared verbatim with the solid track — one experiment serves both"),
 ("I","Liquid","Film cell prototype — THE GATE",150,250,"weeks",1,
  "Outlet RH vs irrigation rate FIRST; 1/2/3-stage NTU/m; face velocity incl. the ~123 m³/h point; rocking rig 5–20°; flooded run",
  "The liquid gate: if the capacity model holds, the bank build proceeds"),
 ("J","Shared","CO₂ sorbent (amine)",150,300,"weeks",1,
  "Capacity at 1,000–1,500 ppm / 45–55% RH; desorption completeness 85–95 °C; water co-adsorption; amine/ammonia slip; fade over 10²–10³ cycles",
  "Required-PENDING: gates X10 entirely and the breathing-air connection. Fallback = oversized ERV + DCV"),
 ("J-K","Shared","CO₂ sorbent (potash variant)",40,80,"weeks",0,
  "K₂CO₃-on-apolar-carbon: capacity, regeneration at 120/135/150 °C with logged purge, alumina-deactivation control, caking, alkaline carryover, 10² cycles",
  "Waste-heat platforms with a ≥~130 °C tap only (X11)"),
 ("L","Upgrade","Hot-film absorption (AHT)",60,120,"weeks",0,
  "Absorption rate and approach of a 43–44 wt% film at 85/88/90 °C against ~20–25 kPa steam; drain-back-on-stop and a deliberate cool-in-place fault",
  "Gates X12. Go/no-go: measurable absorption at 88–90 °C with test A's aw table confirming the margin"),
 ("T1","Solid","Mill-zero synthesis batch",N,N,"days",1,
  "Mortar-and-pestle ~25 g LAG batch with the F4 feedstock branch pre-loaded",
  "Stage 0 of the staged pipeline"),
 ("T2","Solid","Full airflow–moisture transient model",N,N,"weeks",1,
  "Supersedes the doc 00 §4 steady state; sets charge, coated area, condenser duty, fan power",
  "Gates every downstream solid size — the first modelling task before hardware"),
 ("T3","Solid","RH / DAQ channel budget",N,N,"days",1,
  "Wet/dry-bulb + K-type channel plan before M2 (second logger or unified ESP32 node)",
  "On the critical path for M2"),
 ("T4","Solid","Outsourced PXRD + DVS",N,N,"weeks",1,
  "Phase match to A520 and the Type V step at 25–30% RH",
  "The real pass/fail gate on synthesis (QC gates 3 and 5)"),
 ("T5","Solid","Coating SOP",N,N,"weeks",1,
  "PVA grade locked, anneal mandatory, adhesion and capacity retention",
  "Runs on the commercial A520 lot so coating never blocks on synthesis"),
 ("T6","Solid","Bench M1–M4",N,N,"weeks",1,
  "M2 against a representative ~24 g/kg purge → effective Δq",
  "Gates the heat-source decision and the inventory"),
 ("T8","Solid","Silica-wheel benchmark quote",N,N,"days",1,
  "Commercial wheel sized to the T2 duty",
  "The fallback at any gate — and the honest benchmark"),
 ("T9","Solid","Commercial Basolite A520 lot",N,N,"days",1,
  "25–100 g buy-to-validate lot",
  "Parallel path: coating work proceeds independently of synthesis"),
 ("M1","Solid","Adsorption working capacity + coated-face ΔP",N,N,"days",1,
  "Weigh activated-dry coupon → saturate isothermally → reweigh; U-tube manometer at design face velocity",
  "ΔP gates the 0.6–1.0 kW fan-power line"),
 ("M2","Solid","Regeneration vs representative humid purge",N,N,"weeks",1,
  "One 45–50 °C confirmation, then completeness and kinetics at 60/65/70 °C against a logged ~24 g/kg purge",
  "THE most economically consequential measurement: confirms F2, extracts the F3 effective Δq"),
 ("M3","Solid","Specific regeneration energy",N,N,"weeks",1,
  "Fluid-side balance Q = ∫ṁ·c_p·ΔT dt, dry blank first, then wet; condensate cross-check",
  "vs the 0.67 kWh/L latent floor and 0.8–1.0 kWh/L total"),
 ("M4","Solid","Cycling durability + salt-air edge behaviour",N,N,"months",1,
  "10²–10³ swings, periodic M1 + dry-mass weigh, coating edge/defect inspection under salt aerosol",
  "Capacity fade and coating mass loss — the failure mode that matters (F5)"),
]

# =============================================================== FINDINGS ====
FINDINGS = [
 ("F1","Solid","M-cycle working-air moisture is the dominant latent term",
  "Peak duty ~9–11 kg/h (not ~2); regen ~7.5–9 kW; condenser and fan duty scale with it. Closure: condensate covers M-cycle feed with ~50–70 L/day surplus",
  "PENDING T2 (full transient sets the final charge)","00 §4; 20 §5"),
 ("F2","Solid","Solar-direct regeneration is equilibrium-dead at the peak point",
  "Against a condensing purge (~25 g/kg) a 45–50 °C bed faces 29–38% RH at its face — at/above AlFu's step. Waste heat primary; the liquid track is the solar comfort island",
  "PENDING M2 (confirms; real deliverable is 60–65 °C completeness + effective Δq)","20 §6"),
 ("F3","Solid","The sizing Δq is optimistic",
  "0.2–0.3 g/g is a dry-purge full-swing figure; a 0.05–0.10 g/g residual cuts effective Δq to 0.15–0.2 (+25–50% inventory), compounding F1",
  "PENDING M2","21 §6; 22 §4"),
 ("F4","Solid","Mill-zero needs a feedstock-reactivity branch",
  "The literature mortar-and-pestle result used freshly precipitated amorphous Al(OH)₃; crystalline gibbsite is sluggish. A PXRD fail may be feedstock, not milling energy",
  "PENDING T1/T4","21 §4"),
 ("F5","Solid","Aluminium DCHX in chloride service",
  "Sealed/filtered intake, no dissimilar-metal fittings, coating as barrier layer, materials law imported; X8 closure is the primary mitigation",
  "Corrosion rate PENDING M4","22 §3"),
 ("F6","Solid","The DCHX sensible-cycling bucket is unverified",
  "Swung heat capacity per unit water swung is unspecified; if the allowance is optimistic, specific energy and continuous regeneration both move upward and bed-to-bed recovery becomes a hard requirement",
  "PENDING M3 (M1 gates the coated-area denominator)","doc 40 F6; doc 20 §3–4; doc 22 §1"),
 ("X1","Shared","Once-through ventilation and whole-cabin M-cycle never compose",
  "5.5–11 kg/h absorber duty at DP-A; in recirc topology whole-cabin liquid AC converges to solid-track duty (~10.6 kg/h, ~11 kW) — deferred",
  "settled","00 §4/§6; 10 §2"),
 ("X2","Shared","Regeneration asymmetry is a theorem",
  "The sealed still (no purge) keeps positive driving force at any pool >~40 °C; the solid bed hits F2's wall below ~50 °C → liquid = solar layer, solid = waste-heat layer",
  "settled","00 §3; 30 §2"),
 ("X3","Shared","Liquid desiccant for cabin air is re-scoped, not rejected",
  "'Rejected absent demonstrated aerosol control' — test B is the arbiter",
  "PENDING B","21 §2; 11 §3"),
 ("X4","Shared","Materials doctrine flows both ways",
  "Nylon ban, brazed-plate prohibition, drip discipline and the sourcing heuristic imported to the solid track",
  "settled","22 §3"),
 ("X5","Shared","The airflow cascade",
  "Route makeup through the cabin before it becomes working air. Berth scale: full 48 m³/h kept (1,920 vs 2,670 ppm), more cooling (102 vs 86 W), 1.6 °C supply cost",
  "PENDING H","10 §2; 20 §2"),
 ("X6","Liquid","Mixed-mode recirculation is the occupied baseline",
  "Once-through fails 4-adult comfort at DP-A (59–66% RH); mixed-mode restores 55% at 0.88 kg/h. Recirc-only prohibited while occupied",
  "PENDING I","10 §2; 00 §5"),
 ("X7","Shared","ERV latent recovery on the fresh stream",
  "(1−ε)×10.4 g/kg per kg; ε ≥0.8 saves ~9–10 kWh/day and makes generous ventilation cheap. Needs one ducted exhaust path; crossover <5%",
  "PENDING E","11 §5; 20 §2"),
 ("X8","Shared","Exhaust-vapour recovery doctrine",
  "Saturated process exhausts end in distillation recovery; no stream is both rejection sink and recovery source; closed working loop = water-neutral M-cycle",
  "design intent; PENDING H/G","00 §7; 20 §8; 11 §4"),
 ("X9","Shared","All-electric galley — no combustion in the envelope",
  "One 2 kW burner = 3.6× the crew's CO₂ + ~0.25 kg/h latent; induction costs ~2 kWh_e/day",
  "settled (safety register 2)","00 §5; 30 §4"),
 ("X10","Shared","The CO₂ battery",
  "Two-bed solid-amine TSA, 85–95 °C regen, ~1.6 kg/day for ~1.6–2.5 kWh/day — the only path holding <1,000 ppm sealed",
  "REQUIRED-PENDING J","00 §5; 11 §5; 12 §4"),
 ("X11","Shared","Sealed-mode CO₂-bed chemistry is heat-grade-dependent",
  "K₂CO₃/apolar-carbon TSA (130–150 °C) is the waste-heat-platform primary; the amine resin (85–95 °C) is the solar-grade bed; alumina/MgO supports prohibited",
  "REQUIRED-PENDING J-K","00 §5; 12 §4–6; 30 §2/§6"),
 ("X12","Upgrade","The brine's activity depression is a temperature lift",
  "A single-stage CaCl₂ absorption heat transformer — sealed still unchanged as generator + condenser — upgrades a 60–65 °C tail to 85–90 °C at COP 0.45–0.48. Ceiling ≲91 °C single-stage",
  "PENDING L — upgrade path only, never load-bearing","31 §2; 40 §2"),
 ("X14","Shared","Thermal-swing sensible penalty is set by the inert-mass ratio, not by cycle time",
  "Half-cycle duration cancels; the design levers are inert-mass-to-sorbent ratio and bed-to-bed recovery effectiveness, never cycle speed",
  "settled","doc 40 X14; doc 20 §3–4; doc 22 §1"),
 ("P17","Shared","CO₂ specification (safety-critical)",
  "<1,000 ppm at all times in all modes; alarm + forced boost at 2,000 ppm; per-room maximum sensing; mechanical minimum stop on the fresh damper",
  "binding requirement","00 §5/§8"),
]

SAFETY = [
 (1,"CO₂ interlock","<1,000 ppm target · 2,000 ppm alarm/boost · per-room maximum sensing · mechanical minimum stop on the fresh damper","00 §5; P17"),
 (2,"No combustion in the envelope","No gas or combustion appliances in the conditioned envelope (X9)","00 §5; 30 §4"),
 (3,"Aerosol control before cabin connection","Drift eliminator + demister chain and a clean steel-coupon acceptance (test B) before any liquid contactor touches breathing air","11 §3"),
 (4,"CO₂-sorbent breathing-air assay","Amine/ammonia slip (test J) for amine beds; alkaline particulate/mist carryover (test J-K) for carbonate beds","00 §5; 12 §4"),
 (5,"Distillate potability","Independent water-quality test before regular consumption; air-swept-condenser water never enters potable or M-cycle service before a TDS assay clears it","11 §4; 12 §4"),
 (6,"Crystallization interlock","42 wt% high-SG cutoff unless brine ≥22 °C guaranteed; 43 wt% storage cap; drain-back-to-still on stop where an AHT absorber is fitted","11 §4; 31 §2.6"),
 (7,"Two-worlds materials rule","Applied to every brine- or raw-water-wetted component; never relaxed for land installs — the desiccant itself is the chloride source","11 §1; 00 §1"),
]

MATERIALS = [
 ("Use freely","PVC (≤60 °C), CPVC (≤93 °C), PP, HDPE/PE/PE-RT, PEX, PVDF, titanium, EPDM, FKM, silicone, ceramic/alumina, glass, acetal (POM)","Brine and raw-water side is metal-free by rule"),
 ("Never","Copper, brass, bronze; aluminium (incl. anodized); zinc/galvanized; mild steel; 304 SS; nylon/polyamide","CaCl₂ is a documented stress-cracking agent for loaded PA parts (erratum 4)"),
 ("Marginal — do not design in","316/316L","Brief splash only; never warm immersion or crevice geometry (threads, brazed channels, gasket lands)"),
 ("Wetted fasteners","PP / PVC / PVDF only","Parts sold for pools/aquaculture/chemical dosing are probably right; plumbing parts hide brass"),
 ("Prohibited assembly","Any brazed-plate exchanger (304 + copper braze) on brine or raw water","Glycol↔buffer isolation duty only"),
 ("Membrane between worlds","Titanium heat exchangers only","Never let the two fluids mix"),
 ("AHT absorber (85–90 °C brine)","PVDF or PP-with-margin; titanium extraction coil","Past comfortable CPVC territory (93 °C with no margin)"),
 ("Solid-track loop","All-aluminium + stainless wetted path, distilled water, silicone tubing on the hot loop","Avoid brass — galvanic couple with the anodic Al plate"),
]

# consistency observations surfaced while compiling this register
CHECKS = [
 ("CHK-001","Unit label","12 §1","Mixed-mode total absorber flow is given as '~123–147 m³/h'",
  "123 m³/h and 147 kg/h are the same quantity at ρ ≈ 1.2 kg/m³ (147/1.2 = 122.5); the removal build-up 58 kg/h fresh + 90.6 kg/h recirc = 148.6 kg/h confirms it",
  "Low — presentational","RESOLVED v1.3 — doc 12 §1 restated as one quantity in two units"),
 ("CHK-002","Internal consistency","10 §3 vs 12 §1","ERV pre-dried fresh air is 17.4 g/kg in doc 10 §3 and 15.9 g/kg in doc 12 §1",
  "15.9 = 24.2 − 0.8(24.2−13.8) at the specified ε 0.8; 17.4 corresponds to ε ≈ 0.65. The quoted ERV'd removal 0.49 kg/h reproduces only from 17.4 (ε 0.65); ε 0.8 gives 0.40 kg/h",
  "Medium — affects the ERV'd duty line","OPEN — recorded as doc 12 §2 erratum 10 and in doc 40's make-or-break list; gated on test E"),
 ("CHK-003","Basis label","00 §5","'0.06 kWh/day per m³/h' (ε 0.8) and 'each 12 m³/h saves ~3.8 kWh/day' sit in the same paragraph",
  "Both are correct but on different bases: 3.8/12 = 0.317 kWh/day·m³/h is the BARE-fresh figure; 0.06 is the ERV'd one. Ratio 5.3 ≈ 1/(1−ε)",
  "Low — presentational","RESOLVED v1.2 — doc 00 §5 now labels both bases"),
 ("CHK-004","Basis label","10 §3 vs 12 §1","Air-side heat removal quoted as ~35 W/K (doc 10) and ~16 W/K (doc 12)",
  "16.2 W/K is 48 m³/h of fresh air; ~41 W/K would be the 123 m³/h mixed-mode flow. Both are ṁ·c_p at their own flow — not contradictory, but the flow basis is unstated in doc 10",
  "Low — presentational","RESOLVED v1.1 — doc 10 §3 now names the flow"),
 ("CHK-005","Budget arithmetic","README; 12 §4","The '~$485–755 gates everything' headline does not decompose from the doc 12 §4 cost column",
  "The plausible baseline set {A,B,D,E,F,G,H,I,J} sums to $565–885; dropping J gives $415–585. No subset reproduces 485–755 exactly (the range span, 270, matches no combination of the listed spans)",
  "Medium — traceability","RESOLVED v1.4 — the aggregate headline is retired; per-test estimates retained and scoped as 2026 order-of-magnitude figures (doc 12 §4)"),
 ("CHK-006","Model reproduction","00 §4; 20 §5","Independent re-solve of the doc 00 §4 steady state",
  "Reproduces the published band on both topologies (see the Model 00§4 sheet): cabin-draw gives T_sup 18.2 °C, S 0.508 kg/s, ω_cab 9.09 g/kg, M 0.142 kg/s, duty 9.8 kg/h; dry-draw gives T_sup 16.8 °C, S 0.425 kg/s, duty 9.6 kg/h. Both sit inside the published 9–11 kg/h, and regeneration heat lands at 8.2 / 8.3 kW against the published 7.5–9 kW",
  "Informational — model confirmed","The published supply-temperature band 16.8–18.1 °C is exactly the span between the two topologies, which is worth stating explicitly in doc 20 §5"),
 ("CHK-007","Model reproduction","12 §1","Independent re-derivation of the 0.88 kg/h mixed-mode peak",
  "58 kg/h × (24.2−11.9) + 90.6 kg/h × (13.8−11.9) = 0.713 + 0.172 = 0.885 kg/h — exact",
  "Informational — confirmed",""),
 ("CHK-008","Model reproduction","00 §5","Independent re-derivation of the CO₂ dose–response ladder",
  "c = 400 + 10⁶·V̇_CO2/V̇_fresh with V̇_CO2 = 0.072 m³/h reproduces 1,920 / 2,420 / 3,420 / 6,420 ppm and 1,400 ppm asleep exactly — confirming 0.072 m³/h is the four-occupant total, not per person",
  "Informational — confirmed","RESOLVED v1.2 — doc 00 §5 labels the rate 'crew total', foreclosing a 4× misreading"),
 ("CHK-009","Basis constant","12 §1","Air mass flows imply a density of ~1.208 kg/m³",
  "48 m³/h → 58 kg/h and 75 → 90.6 kg/h both back-solve to ρ = 1.208 kg/m³ (standard ~1.2 air, roughly 20 °C). Moist air at DP-A is ~1.145 kg/m³, so the published mass flows — and every duty derived from them — run ~5% high",
  "Low — conservative direction","RESOLVED v1.2 — doc 00 §2 states the ρ 1.2 basis"),
 ("CHK-010","Stated value","20 §5","Cabin steady state given as '~9.1 g/kg / ~40% RH'",
  "9.1 g/kg at the 25 °C cabin target is ~46% RH (Pv 1.461 kPa against Psat 3.169 kPa). ~40% RH at 25 °C would be ω ≈ 7.9 g/kg",
  "Low — no design consequence","RESOLVED v1.1 — doc 20 restated at ~46% RH throughout"),
 ("CHK-011","Model reproduction","20 §8","X8 closure penalty at DP-A",
  "Re-drying saturated exhaust (25.6 g/kg) instead of ambient makeup (24.2 g/kg) raises the working term by ~8% and total duty by ~7% against the same-topology open cycle — consistent with doc 20 §8's '+16% swing, partly bought back'",
  "Informational — confirmed","Confirms the doctrine's core claim: at DP-A ambient is nearly as wet as saturated exhaust, so closing the loop is nearly free"),
 ("CHK-012","Basis label","12 §1; 20 §6","Two figures quote a basis their arithmetic does not use",
  "(a) The 0.92 kW peak regeneration heat needs COP 0.65 — the WITH-recovery-HX value — not the 0.59 build-up printed beside it (0.599/0.65 = 0.921). (b) Doc 20 §6's purge-RH column back-solves to ~23 g/kg, while the text beside it says '~25 g/kg' and M2 is specified at ~24 g/kg",
  "Low — presentational","RESOLVED v1.3/v1.1 — doc 12 §1 names COP 0.65; doc 20 §6 records the table's purge basis"),
 ("CHK-013","Internal consistency","22 §2 vs 22 §4","Coating loading and the sizing table's coated area do not reconcile",
  "Doc 22 §2 specifies ~0.18 kg sorbent/m² planform. The §4 envelope-only row pairs 3–4 kg with 10–13 m², which implies ~0.3 kg/m². At the stated 0.18 kg/m² the same inventory needs ~19 m²",
  "Medium — sizing consequence","OPEN — doc 22 §4 open item and doc 40's make-or-break list; gated on M1, size from 0.18 kg/m² meanwhile"),
 ("CHK-014","Rounding","20 §5","Condensate quoted as 150–220 L/day against a 9–11 kg/h duty",
  "Duty × 24 h gives 216–264 L/day desorbed, so the published condensate range implies that part of the moisture leaves with the ventilation exhaust rather than reaching the condenser — which is physically right but never stated",
  "Low — presentational","RESOLVED v1.1 — doc 20 §5 separates the three"),
 ("CHK-015","Basis mismatch","20 §3","Heat-recovery fraction quoted two ways in one sentence",
  "'Only ~10–15% is recoverable as useful energy' sits beside 'capped by DHW demand (~1–3 kWh/day against 35+ kWh/day liberated)'. The DHW cap is ~3–8% of 35 kWh/day, and only ~1% of the ~200 kWh/day of regeneration input the corrected F1 duty actually implies. The 35 kWh/day figure also predates the F1 correction",
  "Low — no design consequence","RESOLVED v1.1 — doc 20 §3 restated against the corrected F1 duty"),
 ("CHK-016","Sizing basis","00 §5","CO₂-battery duty of ~1.6 kg/day is a daily average, not the peak",
  "Steady-state balance at the 48 m³/h floor: crew 3.42 kg/day all-awake, ventilation removes 1.37 kg/day at 1,000 ppm, so the bed must take 2.05 kg/day — 28% above the published 1.6. The 1.6 figure reproduces only on a 16 h awake / 8 h asleep average. The doc's own 91 m³/h at 50% single-pass capture independently implies 2.16 kg/day, agreeing with the peak rather than with 1.6",
  "HIGH — safety-critical spec","OPEN — P17 requires <1,000 ppm AT ALL TIMES, so bed mass and regeneration heat should be sized on the all-awake peak: restate as ~2.05 kg/day peak (1.6 average), scale the bed ~28%, and have test J measure working capacity at the peak rate"),
 ("CHK-017","Stated threshold","00 §3","X2's '~40 °C' still-crossover is the dilute-end figure only",
  "aw·Psat(T_pool) > Psat(29 °C) gives a crossover of 39.8 °C at aw 0.55 (35 wt%, where regeneration starts), 43.6 °C at aw 0.45 (40 wt%), and 49.1 °C at aw 0.34 (43–44 wt% hot-regen). The single '~40 °C' figure holds only at the start of the concentration swing",
  "Low — no design consequence","OPEN — the design pool is 60–93 °C, far above all three, so X2's conclusion is untouched. State the crossover as a function of target concentration, since it sets the minimum useful pool temperature in degraded mode"),
 ("CHK-018","Cross-track basis","12 §1 vs 20 §1","The two tracks assume latent gains differing ~7× for the same platform",
  "Liquid uses 280 g/h (4 × 70 g/h occupants only); solid uses 1.8–2.8 kg/h 'envelope + occupant'. The 1.72 kg/h non-occupant remainder implies ~0.95 ACH on the solid cabin (~1.37 ACH on the liquid one), against the 0.05–0.15 ACH doc 12 §3 uses for the unattended liquid case and the 0.48 ACH mechanical ventilation floor. Neither track states an ACH",
  "Medium — traceability","OPEN — bounded, not architecture-breaking: re-solving doc 00 §4 across gains of 0.64–2.8 kg/h moves duty only 8.8→10.3 kg/h, so F1 and the 9–11 kg/h band both survive. State one envelope-leakage assumption in doc 00 and let both tracks inherit it"),
 ("CHK-019","Internal consistency","11 §2","Cell rating, face-velocity band and NTU approach share no operating point",
  "A 600×300×150 mm cell has a ~0.09 m² face. At the DP-A baseline (123 m³/h over a 2-cell bank) face velocity is ~0.19 m/s — well under the stated 0.6–1.0 m/s — while the 0.4–0.8 kg/h rating reproduces exactly at that baseline. At 0.6 m/s the same cell passes ~194 m³/h and, at the NTU-1.9 / 85% approach used to size contactor depth, would remove ~2.9 kg/h, far above its own rating",
  "Medium — gated by test I","OPEN — reconcilable only at an approach efficiency that falls with face velocity, which is never stated. Test I already measures exactly this (outlet RH vs irrigation and face velocity, staged NTU/m). Publish the rating with its approach and flow, not as a bare kg/h"),
]

DOCS = [
 ("00_platform_basis.md","v1.3","Scope, DP-A, shared physics, airflow–moisture model, CO₂ stack, X8 doctrine, safety register"),
 ("10_liquid_concept_physics.md","v1.1","Brine principle, mixed-mode baseline, moisture battery, berth cascade, performance envelope"),
 ("11_liquid_architecture_materials.md","v1.2","Two-worlds law, film-cell bank, aerosol control, sealed still, ERV/CO₂ hardware, thermal bus"),
 ("12_liquid_numbers_test_plan.md","v1.4","Validated quantities at DP-A, errata trail, sensitivities, tests A–L, rejected CO₂ alternatives"),
 ("20_solid_concept_system.md","v1.2","Architecture, corrected F1 balance, regeneration-vs-purge physics, X8 closed loop, energy verdicts"),
 ("21_solid_sorbent_synthesis.md","v1.0","Isotherm-step selection, candidates, aqueous + LAG routes, F4 branch, QC gates"),
 ("22_solid_module_validation.md","v1.2","DCHX design, coating rules, F5 mitigations, bench rig, M1–M4, staged pipeline"),
 ("30_integration_energy_water.md","v1.3","Heat cascade, HDH, source roles, all-electric galley, water ladder, degraded operation"),
 ("31_upgrade_paths_sorption_cycles.md","v1.0","X12 AHT, coupled VC heat pump, still MVR, closed AlFu chiller, static crystallizer"),
 ("40_findings_register.md","v1.5","F1–F6, X1–X12, X14, spec P17, tasks, make-or-break bench list"),
 ("50_defensive_disclosure_plan.md","v1.3","Venue stack, repo formation, Zenodo procedure, metadata, version discipline"),
 ("executive_summary.md","v1.1","Standalone abstract for examiner-channel deposit"),
]


# ====================================================== FORMULA FRAGMENTS ====
# Every fragment references DEFINED NAMES, never a literal constant.
def f_psat(t):       return f"(MAG_A*EXP(MAG_B*{t}/({t}+MAG_C)))"
def f_wsat(t):       return f"(RATIO_MW*{f_psat(t)}/(P_atm-{f_psat(t)}))"
def f_w(t, rh):      return f"(RATIO_MW*{rh}*{f_psat(t)}/(P_atm-{rh}*{f_psat(t)}))"
def f_pv(w):         return f"(P_atm*{w}/(RATIO_MW+{w}))"
def f_tdp(w):        return f"(MAG_C*LN({f_pv(w)}/MAG_A)/(MAG_B-LN({f_pv(w)}/MAG_A)))"
def f_tsat_at(p):    return f"(MAG_C*LN({p}/MAG_A)/(MAG_B-LN({p}/MAG_A)))"
def f_h(t, w):       return f"(CP_A*{t}+{w}/1000*(H_FG0+CP_V*{t}))"


def put(ws, r, c, v, style="s_txt", fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.style = style
    if fmt:
        cell.number_format = fmt
    return cell


def kv(ws, r, label, value, unit="", note="", style="s_in", fmt="0.###"):
    put(ws, r, 1, label, "s_lbl")
    put(ws, r, 2, value, style, fmt)
    put(ws, r, 3, unit)
    put(ws, r, 4, note, "s_note")


def delta_row(ws, r, label, calc_formula, published, unit, note, tol=0.10):
    """calculated | published | relative delta | verdict"""
    put(ws, r, 1, label, "s_lbl")
    put(ws, r, 2, calc_formula, "s_calc", "0.000")
    put(ws, r, 3, published, "s_ref", "0.000")
    put(ws, r, 4, f"=IF(C{r}=0,\"\",(B{r}-C{r})/C{r})", "s_calc", "0.0%")
    put(ws, r, 5, f'=IF(C{r}=0,"—",IF(ABS(D{r})<={tol},"OK","REVIEW"))', "s_calc")
    put(ws, r, 6, unit)
    put(ws, r, 7, note, "s_note")
    ws.conditional_formatting.add(f"E{r}", CellIsRule(
        operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=OKG),
        font=Font(color=OKT, bold=True)))
    ws.conditional_formatting.add(f"E{r}", CellIsRule(
        operator="equal", formula=['"REVIEW"'], fill=PatternFill("solid", fgColor=WARN),
        font=Font(color=WARNT, bold=True)))


# ================================================================ SHEETS =====
def sheet_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_properties.tabColor = NAVY
    for col, w in zip("ABCDEF", (26, 24, 20, 20, 20, 20)):
        ws.column_dimensions[col].width = w
    put(ws, 1, 1, "Heat-Driven Comfort & Water — Parameter Register", "s_title")
    ws.merge_cells("A1:F1"); ws.row_dimensions[1].height = 24
    put(ws, 2, 1, "Every quantitative claim in the document lineage, in one traceable "
                  "workbook, with live re-derivations of the headline numbers.", "s_note")
    ws.merge_cells("A2:F2")
    rows = [
        ("Workbook revision", REV),
        ("Source document set", DOCSET),
        ("Repository", REPO),
        ("Concept DOI (all versions)", CONCEPT_DOI),
        ("Version DOI (v1.0 record)", VERSION_DOI),
        ("Governing design point", "DP-A · 32 °C / 80% RH · ω 24.2 g/kg · dew point 28.1 °C · raw-water sink ~29 °C"),
        ("Status of the work", "Paper design. NOTHING HAS BEEN BUILT. No figure here is a measurement "
                               "unless its grade says 'measured'."),
        ("Licensing", "Hardware CERN-OHL-P v2 · documentation & this workbook CC-BY-4.0 · "
                      "generator script MIT (LICENSE.md)"),
        ("Disclosure", "Open defensive publication. No patents sought or held."),
        ("Generated by", "scripts/build_parameter_workbook.py — regenerate rather than hand-edit"),
    ]
    r = 4
    for k, v in rows:
        put(ws, r, 1, k, "s_lbl"); put(ws, r, 2, v)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    r += 1
    band(ws, r, "SAFETY — binding on any build (doc 00 §8)", 6); r += 1
    put(ws, r, 1, "CO₂ interlock (<1,000 ppm target / 2,000 alarm / per-room maximum sensing / "
                  "mechanical minimum stop) · no combustion appliances in the envelope · aerosol and "
                  "slip assays before any cabin connection · potability and TDS tests before any water "
                  "is drunk · crystallization interlock · two-worlds materials rule. "
                  "Any build omitting these departs from this design.", "s_txt")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
    ws.row_dimensions[r].height = 30
    r += 4

    band(ws, r, "SHEET INDEX", 6); r += 1
    put(ws, r, 1, "Sheet", "s_hdr"); put(ws, r, 2, "Contents", "s_hdr")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6); r += 1
    for name, desc in [
        ("Legend", "Confidence grades, column definitions, cell conventions, unit policy"),
        ("Design Point", "The named constants every calculation sheet references"),
        ("Register", "Master parameter register — filterable, one row per published value"),
        ("Psychrometrics", "Magnus formulation; converter; saturation table validated against doc 00 §2"),
        ("Model 00§4", "The generalized airflow–moisture model, solved live for three topologies"),
        ("Liquid Sizing", "Mixed-mode removal, regeneration, reserve and rejection — live from DP-A"),
        ("Solid Sizing", "Duty → inventory → coated area → condensate balance, with the F3 penalty"),
        ("CO2 & Ventilation", "Dose–response ladder, ERV economics, scrubber duty — live"),
        ("Heat & Water", "Daily heat budget, collector/fuel sizing, water redundancy ladder"),
        ("Upgrade Paths", "X12 lift-ceiling inequality and the MVR pressure ratio, solved live"),
        ("Test Program", "Tests A–L and T/M series with costs, durations, and the live budget sum"),
        ("Findings", "F1–F6, X1–X12, X14, spec P17 — status and where each lives"),
        ("Safety & Materials", "The binding register and the two-worlds materials law"),
        ("Checks", "Consistency observations raised while compiling this register"),
        ("Sources", "Document set, versions, and scope"),
    ]:
        put(ws, r, 1, name, "s_lbl"); put(ws, r, 2, desc)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6); r += 1
    return ws


def sheet_legend(wb):
    ws = wb.create_sheet("Legend")
    titleblock(ws, "Legend & conventions", "How to read — and how to extend — this register.", 5)
    header(ws, [("Confidence grade", 22), ("Meaning", 60), ("Rule", 62), ("", 2), ("", 2)], row=4)
    grades = [
        ("measured", "A number this project has physically measured", "None yet — the program is unbuilt"),
        ("measured (lit.)", "Measured by others and cited", "Cite the regime; literature values rarely transfer across purge humidity"),
        ("procurement-grade", "From a datasheet, COA, or supplier quote", "Verify per lot; hydrate assays drift"),
        ("sizing-grade", "Derived from first principles or published correlations", "Usable for sizing, never for acceptance"),
        ("estimate-grade", "Doc 31 upgrade paths — no bench basis at all", "Never load-bearing for any DP-A claim"),
        ("PENDING <test>", "Explicitly gated on a named experiment", "Do not upgrade the grade without the result"),
        ("requirement", "A specification the design must meet", "Binding; changes require a version bump"),
        ("governing / platform", "Fixed by the design point or the platform definition", "Changing one invalidates every derived number"),
        ("rejected / superseded", "Retained as the record of what was tried", "Kept visible on purpose — the correction trail"),
    ]
    r = 5
    for g, m, rule in grades:
        put(ws, r, 1, g, "s_lbl"); put(ws, r, 2, m); put(ws, r, 3, rule); r += 1
    r += 1
    band(ws, r, "CELL CONVENTIONS", 5); r += 1
    for a, b in [
        ("Blue bold on pale fill", "INPUT — safe to change; every calculation downstream follows"),
        ("Plain right-aligned", "CALCULATED — a live formula; do not overwrite"),
        ("Grey italic", "PUBLISHED value quoted from the documents for comparison"),
        ("OK / REVIEW", "Automatic check: calculated vs published within ±10%"),
        ("VALUE vs MIN/MAX", "A single-point figure goes in VALUE; a published band goes in MIN and MAX"),
        ("Units", "SI throughout; humidity ratio in g/kg dry air, heat in kW and kWh/day, "
                  "concentration in wt%, CO₂ in ppm by volume"),
        ("Traceability", "Every register row names its source document and section. "
                         "If a row has no source, it does not belong here"),
        ("Recalculation", "The workbook is saved with full-calculation-on-load; formulas re-solve when opened"),
    ]:
        put(ws, r, 1, a, "s_lbl"); put(ws, r, 2, b)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); r += 1
    return ws


def sheet_designpoint(wb):
    """The named-constant sheet. Every other sheet references these by name."""
    ws = wb.create_sheet("Design Point")
    ws.sheet_properties.tabColor = STEEL
    titleblock(ws, "Design point & named constants",
               "DP-A is the sole design point — the continuous-duty maximum for every mode. "
               "Change a blue cell and every calculation sheet follows.", 5)
    for col, w in zip("ABCDE", (34, 14, 14, 74, 4)):
        ws.column_dimensions[col].width = w

    defs = []          # (defined_name, sheet_cell)
    r = 4

    def const(name, label, value, unit, note, style="s_in"):
        nonlocal r
        kv(ws, r, label, value, unit, note, style)
        defs.append((name, f"'Design Point'!$B${r}"))
        r += 1

    band(ws, r, "AMBIENT — DP-A (doc 00 §2)", 4); r += 1
    const("T_amb", "Ambient dry-bulb", 32, "°C", "Sole design point; excursions are transient, never a sizing basis")
    const("RH_amb", "Ambient relative humidity", 0.80, "—", "Entered as a fraction")
    const("T_sink", "Raw-water sink", 29, "°C", "Marine tropical surface; land sinks often cooler = margin")
    const("P_atm", "Atmospheric pressure", 101.325, "kPa", "Magnus basis for every ω conversion")
    const("P_MOIST", "Partial-pressure denominator used in doc 12", 97.5, "kPa",
          "Doc 12 §1 divides by 97.5 kPa (P_atm − Pv) rather than P_atm — reproduced here for traceability")

    r += 1
    band(ws, r, "PLATFORM", 4); r += 1
    const("V_CAB", "Conditioned volume", 100, "m³", "Yacht interior / small dwelling / cabin / shelter")
    const("N_OCC", "Occupants", 4, "person", "")
    const("G_PERS", "Latent generation per occupant", 70, "g/h", "50 sleeping · 90 active; ±30% band")
    const("Q_SENS", "Peak sensible load (solid track)", 3.5, "kW", "Solar noon, ~100 m³ envelope")
    const("V_FRESH", "Fresh-air floor", 48, "m³/h", "12 m³/h·person — CO₂-interlocked, mechanical minimum stop")
    const("CO2_GEN", "CO₂ generation, crew total awake", 0.072, "m³/h", "0.047 asleep; ≈3.4 kg/day")
    const("CO2_OUT", "Outdoor CO₂ baseline", 400, "ppm", "")
    const("CO2_SPEC", "CO₂ specification (P17)", 1000, "ppm", "Safety-critical, all modes including sealed")

    r += 1
    band(ws, r, "COMFORT TARGETS", 4); r += 1
    const("T_CAB_L", "Cabin target, liquid track", 29, "°C", "At 55% RH → ω 13.8 g/kg")
    const("RH_CAB_L", "Cabin target RH, liquid track", 0.55, "—", "")
    const("T_CAB_S", "Cabin target, solid track", 25, "°C", "40–55% RH band")

    r += 1
    band(ws, r, "SORBENT BACK-ENDS", 4); r += 1
    const("AW_40", "Water activity, 40 wt% CaCl₂", 0.45, "—", "PENDING test A — ±0.05 is the largest single unknown")
    const("AW_44", "Water activity, 43–44 wt% CaCl₂", 0.34, "—", "PENDING test A — the hot-regen lever")
    const("T_BRINE", "Brine temperature at the contactor", 30, "°C", "Sink 29 °C + approach; drives the floor band")
    const("C_HI", "Strong-brine concentration", 40, "wt%", "")
    const("C_LO", "Dilute-brine concentration", 35, "wt%", "")
    const("W_SUP_S", "Supply humidity ratio, solid track", 8, "g/kg", "AlFu back-end — the model's only sorbent input")

    r += 1
    band(ws, r, "EFFECTIVENESS & EFFICIENCY", 4); r += 1
    const("EPS_LAT", "ERV latent effectiveness", 0.80, "—", "PENDING test E; ≥0.8 is the X7 basis")
    const("EPS_DP", "M-cycle dew-point effectiveness", 0.70, "—", "0.65–0.80 band; 0.70 is the model value")
    const("COP_REGEN", "Liquid regeneration thermal COP", 0.59, "—", "Build-up in doc 12 §1; 0.65 with a recovery HX")
    const("E_SPEC_S", "Solid specific regeneration energy", 0.85, "kWh/L", "0.8–1.0 band; latent floor is 0.67")

    r += 1
    band(ws, r, "PHYSICAL CONSTANTS (do not edit)", 4); r += 1
    const("MAG_A", "Magnus coefficient A", 0.61094, "kPa", "Psat = A·exp(B·T/(T+C)), T in °C", "s_num")
    const("MAG_B", "Magnus coefficient B", 17.625, "—", "Alduchov–Eskridge form", "s_num")
    const("MAG_C", "Magnus coefficient C", 243.04, "°C", "", "s_num")
    const("RATIO_MW", "Mass ratio water/dry air", 0.622, "—", "M_w/M_a = 18.015/28.964", "s_num")
    const("CP_A", "Specific heat, dry air", 1.006, "kJ/kg·K", "", "s_num")
    const("CP_V", "Specific heat, water vapour", 1.86, "kJ/kg·K", "", "s_num")
    const("H_FG0", "Latent heat at 0 °C", 2501, "kJ/kg", "Enthalpy datum for the moist-air chain", "s_num")
    const("H_FG_REGEN", "Latent heat at regeneration", 2440, "kJ/kg", "2.44 MJ/kg — the doc 10 §3 basis", "s_num")
    const("H_ABS", "Absorption heat into brine", 2700, "kJ/kg", "2.7 MJ/kg — exceeds latent by the heat of dilution", "s_num")
    const("RHO_A", "Air density at DP-A (physical)", 1.164, "kg/m³",
          "Dry air at 32 °C, 101.325 kPa. Moist air at DP-A is lower still (~1.145)", "s_num")
    const("RHO_A_DOC", "Air density basis used in the documents", 1.208, "kg/m³",
          "Back-solved from doc 12's 48 m³/h → 58 kg/h and 75 → 90.6 kg/h (standard ~1.2 air). "
          "Used for the liquid mass flows so the published figures reproduce — see CHK-009", "s_num")
    const("SEC_H", "Seconds per hour", 3600, "s/h", "", "s_num")

    r += 2
    put(ws, r, 1, "DERIVED AT DP-A (live — these reproduce the doc 00 §2 / doc 12 §1 headline figures)", "s_sect")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); r += 1
    for nm, lbl, formula, unit, note in [
        ("PSAT_AMB", "Saturation pressure at ambient", f"={f_psat('T_amb')}", "kPa", "Magnus at 32 °C — doc 12 quotes 4.76"),
        ("PV_AMB", "Ambient vapour pressure", "=RH_amb*PSAT_AMB", "kPa", "Doc 12 quotes 3.81"),
        ("W_AMB", "Ambient humidity ratio", "=1000*RATIO_MW*PV_AMB/(P_atm-PV_AMB)", "g/kg", "Doc 00/12 publish 24.2"),
        ("W_AMB_DOC", "Ambient ω on the doc-12 denominator", "=1000*RATIO_MW*PV_AMB/P_MOIST", "g/kg", "The 0.622·3.81/97.5 form quoted in doc 12 §1"),
        ("T_DP_AMB", "Ambient dew point", f"={f_tdp('W_AMB/1000')}", "°C", "Doc 00 publishes 28.1"),
        ("W_CAB_L", "Cabin ω, liquid target", f"=1000*{f_w('T_CAB_L','RH_CAB_L')}", "g/kg", "Doc 12 publishes 13.8"),
        ("W_SAT_SINK", "Saturation ω at the sink", f"=1000*{f_wsat('T_sink')}", "g/kg", "Doc 00 publishes 25.6 — the working-exhaust state"),
        ("W_ERV", "ERV pre-dried fresh air", "=W_AMB-EPS_LAT*(W_AMB-W_CAB_L)", "g/kg", "Doc 12 publishes 15.9 at ε 0.8 — see Checks CHK-002"),
        ("M_FRESH", "Fresh-air mass flow", "=V_FRESH*RHO_A", "kg/h", "Doc 12 quotes ≈58 kg/h"),
        ("G_LAT", "Cabin latent gains", "=N_OCC*G_PERS", "g/h", "Doc 12 quotes 280 g/h"),
    ]:
        put(ws, r, 1, lbl, "s_lbl"); put(ws, r, 2, formula, "s_calc", "0.00"); put(ws, r, 3, unit)
        put(ws, r, 4, note, "s_note")
        defs.append((nm, f"'Design Point'!$B${r}"))
        r += 1

    for name, ref in defs:
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
    return ws


def sheet_register(wb):
    ws = wb.create_sheet("Register")
    ws.sheet_properties.tabColor = "C00000"
    titleblock(ws, "Master parameter register",
               "One row per published quantitative claim. Filter by track, category, grade, or gating test. "
               "Every row names its source — a row without a source does not belong here.", 13)
    cols = [("ID", 9), ("Track", 10), ("Category", 15), ("Parameter", 42), ("Symbol", 10),
            ("Value", 10), ("Min", 9), ("Max", 9), ("Unit", 14), ("Confidence", 16),
            ("Gate", 10), ("Basis / derivation", 64), ("Source", 14)]
    header(ws, cols, row=4)
    r = 5
    for row in REGISTER:
        for i, v in enumerate(row, start=1):
            st = "s_num" if i in (6, 7, 8) else ("s_lbl" if i == 1 else "s_txt")
            put(ws, r, i, v, st)
        r += 1
    last = r - 1
    ws.auto_filter.ref = f"A4:M{last}"
    ws.freeze_panes = "D5"
    # grade highlighting: anything PENDING or REQUIRED stands out
    ws.conditional_formatting.add(f"J5:J{last}", FormulaRule(
        formula=[f'ISNUMBER(SEARCH("PENDING",$J5))'],
        fill=PatternFill("solid", fgColor=WARN), font=Font(color=WARNT)))
    ws.conditional_formatting.add(f"J5:J{last}", FormulaRule(
        formula=[f'OR($J5="rejected",$J5="superseded",$J5="historical")'],
        fill=PatternFill("solid", fgColor=GREY), font=Font(color=MUTE, italic=True)))
    ws.conditional_formatting.add(f"J5:J{last}", FormulaRule(
        formula=[f'OR($J5="requirement",$J5="governing")'],
        fill=PatternFill("solid", fgColor=OKG), font=Font(color=OKT, bold=True)))

    # summary block above the table
    put(ws, 3, 1, f"Rows: {len(REGISTER)}", "s_lbl")
    put(ws, 3, 3, f'=CONCATENATE("PENDING a named test: ",COUNTIF(J5:J{last},"PENDING*")," rows")', "s_lbl")
    put(ws, 3, 6, f'=CONCATENATE("Requirements: ",COUNTIF(J5:J{last},"requirement")," rows")', "s_lbl")
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
    ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=9)
    return ws


def sheet_psychro(wb):
    ws = wb.create_sheet("Psychrometrics")
    titleblock(ws, "Psychrometrics — Magnus formulation",
               "The shared basis under every number in the repository. The saturation table below "
               "reproduces doc 00 §2; the converter is a working tool for new states.", 8)
    for col, w in zip("ABCDEFGH", (30, 13, 13, 12, 12, 12, 14, 60)):
        ws.column_dimensions[col].width = w
    r = 4
    band(ws, r, "FORMULATION", 8); r += 1
    for a, b in [
        ("Saturation pressure", "P_sat(T) = A·exp(B·T/(T+C)) kPa, with A 0.61094, B 17.625, C 243.04 (Alduchov–Eskridge)"),
        ("Humidity ratio", "ω = 0.622·P_v/(P_atm − P_v), kg/kg — reported here in g/kg dry air"),
        ("Dew point", "T_dp = C·ln(P_v/A)/(B − ln(P_v/A))"),
        ("Moist-air enthalpy", "h = 1.006·T + ω·(2501 + 1.86·T) kJ/kg dry air"),
        ("Why it matters", "At DP-A every ambient sink sits at or above the 28.1 °C dew point — a sink-cooled "
                           "condenser harvests ≈0 water and evaporative cooling cannot pass the dew point. "
                           "Only a desiccant breaks that floor (doc 00 §3)"),
    ]:
        put(ws, r, 1, a, "s_lbl"); put(ws, r, 2, b)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8); r += 1

    r += 1
    band(ws, r, "SATURATION TABLE — validated against doc 00 §2", 8); r += 1
    hdr = r
    for i, t in enumerate(["T (°C)", "P_sat (kPa)", "ω_sat calc (g/kg)", "ω_sat doc 00 (g/kg)",
                           "Δ (%)", "Check", "", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    doc_wsat = {26: 21.3, 27: 22.6, 28: 24.1, 29: 25.6, 30: 27.1}
    for t in range(20, 41):
        put(ws, r, 1, t, "s_in", "0")
        put(ws, r, 2, f"={f_psat(f'A{r}')}", "s_calc", "0.000")
        put(ws, r, 3, f"=1000*RATIO_MW*B{r}/(P_atm-B{r})", "s_calc", "0.00")
        if t in doc_wsat:
            put(ws, r, 4, doc_wsat[t], "s_ref", "0.0")
            put(ws, r, 5, f"=(C{r}-D{r})/D{r}", "s_calc", "0.0%")
            put(ws, r, 6, f'=IF(ABS(E{r})<=0.01,"OK","REVIEW")', "s_calc")
            ws.conditional_formatting.add(f"F{r}", CellIsRule(
                operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=OKG),
                font=Font(color=OKT, bold=True)))
            ws.conditional_formatting.add(f"F{r}", CellIsRule(
                operator="equal", formula=['"REVIEW"'], fill=PatternFill("solid", fgColor=BADF),
                font=Font(color=BADT, bold=True)))
        if t == 29:
            put(ws, r, 8, "Saturated working-air exhaust state in the solid balance", "s_note")
        if t == 28:
            put(ws, r, 8, "≈ the DP-A dew point — the floor every ambient sink sits above", "s_note")
        r += 1

    r += 1
    band(ws, r, "STATE CONVERTER — enter any two of T / RH / ω", 8); r += 1
    hdr = r
    for i, t in enumerate(["State", "T (°C)", "RH (—)", "ω (g/kg)", "P_v (kPa)",
                           "T_dp (°C)", "h (kJ/kg)", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    states = [
        ("Ambient DP-A", 32, 0.80, "Ambient — the governing point"),
        ("Cabin, liquid target", 29, 0.55, "ω 13.8 g/kg published"),
        ("Cabin, solid steady state", 25, 0.40, "ω ≈ 9.1 g/kg published"),
        ("Absorber outlet, base brine", 30, None, "Floor set by sump concentration, not per-pass approach"),
        ("Absorber outlet, hot-regen brine", 30, None, "43–44 wt%, aw 0.34"),
        ("ERV-dried fresh air", 31, None, "Pre-dried from ambient at ε_lat"),
        ("Solid supply air", 17.5, None, "M-cycle product on dried air"),
        ("Working exhaust (saturated)", 29, 1.00, "Recycled in X8 mode, never vented"),
    ]
    for name, t, rh, note in states:
        put(ws, r, 1, name, "s_lbl")
        put(ws, r, 2, t, "s_in", "0.0")
        if rh is not None:
            put(ws, r, 3, rh, "s_in", "0.00")
            put(ws, r, 4, f"=1000*RATIO_MW*C{r}*{f_psat(f'B{r}')}/(P_atm-C{r}*{f_psat(f'B{r}')})", "s_calc", "0.00")
        else:
            # driven by the brine/back-end floor instead of an RH
            if "hot-regen" in name:
                put(ws, r, 4, "=1000*RATIO_MW*AW_44*"+f_psat("T_BRINE")+"/(P_atm-AW_44*"+f_psat("T_BRINE")+")", "s_calc", "0.00")
            elif "base brine" in name:
                put(ws, r, 4, "=1000*RATIO_MW*AW_40*"+f_psat("T_BRINE")+"/(P_atm-AW_40*"+f_psat("T_BRINE")+")", "s_calc", "0.00")
            elif "ERV" in name:
                put(ws, r, 4, "=W_ERV", "s_calc", "0.00")
            else:
                put(ws, r, 4, "=W_SUP_S", "s_calc", "0.00")
            put(ws, r, 3, f"=D{r}/1000*P_atm/(RATIO_MW+D{r}/1000)/{f_psat(f'B{r}')}", "s_calc", "0.00")
        put(ws, r, 5, f"={f_pv(f'D{r}/1000')}", "s_calc", "0.000")
        put(ws, r, 6, f"={f_tdp(f'D{r}/1000')}", "s_calc", "0.0")
        put(ws, r, 7, f"={f_h(f'B{r}', f'D{r}')}", "s_calc", "0.0")
        put(ws, r, 8, note, "s_note")
        r += 1

    r += 1
    band(ws, r, "BRINE FLOOR MATRIX — ω_out = f(aw, brine T)   [PENDING test A]", 8); r += 1
    hdr = r
    put(ws, hdr, 1, "Brine T (°C) ↓ / aw →", "s_hdr")
    aws = [0.33, 0.34, 0.40, 0.45, 0.50]
    for j, a in enumerate(aws, start=2):
        put(ws, hdr, j, a, "s_hdr")
    put(ws, hdr, len(aws) + 2, "Published anchor", "s_hdr")
    r += 1
    for t in (27, 28, 29, 30, 31, 32, 33):
        put(ws, r, 1, t, "s_in", "0")
        for j, a in enumerate(aws, start=2):
            put(ws, r, j, f"=1000*RATIO_MW*{a}*{f_psat(f'$A{r}')}/(P_atm-{a}*{f_psat(f'$A{r}')})", "s_calc", "0.0")
        anchor = {27: "10.0 @ aw 0.45 · 7.5 @ aw 0.34", 30: "11.9 @ aw 0.45 · 9.0 @ aw 0.34",
                  33: "14.2 @ aw 0.45"}.get(t, "")
        put(ws, r, len(aws) + 2, anchor, "s_note")
        r += 1
    put(ws, r, 1, "Published floors are 'sizing-grade, PENDING test A'. Published data carries ±0.05 aw, "
                  "which is ±1.5–3 g/kg on the floor — the largest single unknown in the liquid track "
                  "(doc 12 §3). Cooling leverage 0.5–0.7 g/kg per K of brine cooling is the slope of "
                  "each column.", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=8)
    return ws


def sheet_model(wb):
    """The doc 00 §4 generalized airflow-moisture model, solved live.

    The published system is circular (S depends on T_sup, T_sup on the working-air
    dew point, the working air on Q_wet which depends on S). It is closed here by
    treating the working-air ENTRY STATE as the topology choice and solving in one
    pass, which is exactly how doc 00 §4 describes the closure.
    """
    ws = wb.create_sheet("Model 00§4")
    ws.sheet_properties.tabColor = "7030A0"
    titleblock(ws, "The generalized airflow–moisture model (doc 00 §4)",
               "The closed steady state that replaced the hand-chained state tables — the origin of "
               "finding F1 and the 9–11 kg/h solid-track duty. Solved live for three working-air topologies.", 9)
    for col, w in zip("ABCDEFGHI", (38, 13, 13, 13, 11, 13, 13, 13, 58)):
        ws.column_dimensions[col].width = w
    r = 4
    band(ws, r, "THE SIX EQUATIONS (doc 00 §4)", 9); r += 1
    for eq, txt in [
        ("1", "ω_cab = ω_sup + latent_gains / S                        (cabin steady state)"),
        ("2", "T_sup = T_in − ε_dp·(T_in − T_dp(ω_work))               (M-cycle on dried air)"),
        ("3", "S = Q_sens / (c_p·(T_cab − T_sup))                      (supply flow closes the sensible load)"),
        ("4", "Q_wet = S·c_p·(T_in − T_sup)                            (heat dumped into the wet channel)"),
        ("5", "M = Q_wet / Δh_work,  Δh_work = h_sat(T_exh) − h(entry) (working-air flow)"),
        ("6", "duty = (S − M)(ω_cab − ω_sup) + M(ω_amb − ω_sup)        (sorbent duty)"),
    ]:
        put(ws, r, 1, f"Equation {eq}", "s_lbl"); put(ws, r, 2, txt)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9); r += 1
    put(ws, r, 1, "Why it matters", "s_lbl")
    put(ws, r, 2, "Omitting the M-cycle working-air moisture term under-sizes the desiccant 3–5×. "
                  "At DP-A the 29 °C sink cannot absorb heat from a 25 °C cabin, so ALL cabin sensible "
                  "load leaves evaporatively — and that moisture must pass through the desiccant. "
                  "This is finding F1.", "s_txt")
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=9); r += 3

    band(ws, r, "INPUTS", 9); r += 1
    inp = r
    for lbl, val, unit, note in [
        ("Sensible load Q_sens", "=Q_SENS", "kW", "Doc 20 §1, peak sun"),
        ("Latent gains", "=G_LAT/1000", "kg/h", "Occupants + envelope; doc 20 uses 2.0 kg/h total incl. envelope"),
        ("Envelope + occupant latent (doc 20 basis)", 2.0, "kg/h", "The value doc 20 §5 actually solves with"),
        ("Supply humidity ratio ω_sup", "=W_SUP_S", "g/kg", "AlFu back-end — the model's only sorbent input"),
        ("Intake temperature T_in", 31, "°C", "Post-ERV fresh / mixed state entering the desiccant"),
        ("Cabin temperature T_cab", "=T_CAB_S", "°C", "Solid-track target"),
        ("M-cycle effectiveness ε_dp", "=EPS_DP", "—", ""),
        ("Working exhaust temperature", "=T_sink", "°C", "Saturated at the sink"),
        ("Ambient ω", "=W_AMB", "g/kg", "DP-A"),
    ]:
        kv(ws, r, lbl, val, unit, note, "s_in" if not isinstance(val, str) else "s_calc",
           "0.000")
        r += 1
    Q = f"B{inp}"; GAIN = f"B{inp+2}"; WSUP = f"B{inp+3}"; TIN = f"B{inp+4}"
    TCAB = f"B{inp+5}"; EPS = f"B{inp+6}"; TEXH = f"B{inp+7}"; WAMB = f"B{inp+8}"

    r += 1
    band(ws, r, "SOLUTION BY TOPOLOGY — the working air's entry state is the design choice", 9); r += 1
    hdr = r
    heads = ["Quantity", "Dry-draw\n(post-desiccant)", "Cabin-draw\n(open cycle)",
             "X8 closed loop\n(recycled exhaust)", "Unit", "Published\n(doc 20 §5)", "Δ vs published",
             "Check", "Note"]
    for i, t in enumerate(heads, start=1):
        put(ws, hdr, i, t, "s_hdr")
    ws.row_dimensions[hdr].height = 40
    r += 1

    # working-air entry states per topology
    rows = []
    def mrow(label, fdry, fcab, fx8, unit, published, note, fmt="0.000", tol=0.10):
        nonlocal r
        put(ws, r, 1, label, "s_lbl")
        put(ws, r, 2, fdry, "s_calc", fmt)
        put(ws, r, 3, fcab, "s_calc", fmt)
        put(ws, r, 4, fx8, "s_calc", fmt)
        put(ws, r, 5, unit)
        if published is not None:
            put(ws, r, 6, published, "s_ref", fmt)
            put(ws, r, 7, f"=(C{r}-F{r})/F{r}", "s_calc", "0.0%")
            put(ws, r, 8, f'=IF(ABS(G{r})<={tol},"OK","REVIEW")', "s_calc")
            ws.conditional_formatting.add(f"H{r}", CellIsRule(
                operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=OKG),
                font=Font(color=OKT, bold=True)))
            ws.conditional_formatting.add(f"H{r}", CellIsRule(
                operator="equal", formula=['"REVIEW"'], fill=PatternFill("solid", fgColor=WARN),
                font=Font(color=WARNT, bold=True)))
        put(ws, r, 9, note, "s_note")
        rows.append(r); r += 1
        return rows[-1]

    # entry temperature / humidity of the working air, per topology
    r_went = mrow("Working-air entry ω", f"={WSUP}", 9.1, f"={WSUP}", "g/kg", None,
                  "Dry-draw and X8 re-dry to ω_sup. Cabin-draw carries CABIN humidity, which the model "
                  "also solves — so this cell is a seed; the residual row below reports convergence")
    ws.cell(row=r_went, column=3).style = "s_in"
    ws.cell(row=r_went, column=3).number_format = "0.000"
    r_tent = mrow("Working-air entry T", f"={TIN}", f"={TCAB}", f"={TIN}", "°C", None,
                  "X8 re-dries AND intercools the recycled exhaust exactly like the process stream "
                  "(doc 20 §2), so its entry state matches dry-draw; only the moisture it carries differs")
    r_tdpw = mrow("Working-air dew point", f"={f_tdp(f'B{r_went}/1000')}",
                  f"={f_tdp(f'C{r_went}/1000')}", f"={f_tdp(f'D{r_went}/1000')}", "°C", None,
                  "Equation 2's floor — the M-cycle cannot pass it")
    r_tsup = mrow("Supply temperature T_sup", f"={TIN}-{EPS}*({TIN}-B{r_tdpw})",
                  f"={TIN}-{EPS}*({TIN}-C{r_tdpw})", f"={TIN}-{EPS}*({TIN}-D{r_tdpw})",
                  "°C", 17.45, "Published band 16.8–18.1 °C (midpoint quoted here)", "0.00")
    r_S = mrow("Supply flow S", f"={Q}/(CP_A*({TCAB}-B{r_tsup}))", f"={Q}/(CP_A*({TCAB}-C{r_tsup}))",
               f"={Q}/(CP_A*({TCAB}-D{r_tsup}))", "kg/s", 0.47,
               "Published band 0.43–0.51 kg/s (≈1,350–1,590 m³/h)", "0.000")
    r_Sv = mrow("Supply flow, volumetric", f"=B{r_S}*SEC_H/RHO_A", f"=C{r_S}*SEC_H/RHO_A",
                f"=D{r_S}*SEC_H/RHO_A", "m³/h", 1470, "", "0")
    r_wcab = mrow("Cabin ω (equation 1)", f"={WSUP}+{GAIN}/(B{r_S}*SEC_H)*1000",
                  f"={WSUP}+{GAIN}/(C{r_S}*SEC_H)*1000", f"={WSUP}+{GAIN}/(D{r_S}*SEC_H)*1000",
                  "g/kg", 9.1, "Doc 20 §5: ~9.1 g/kg ≈ 40% RH — drier than target, headroom exists", "0.00")
    r_rh = mrow("Cabin RH at that state", f"=B{r_wcab}/1000*P_atm/(RATIO_MW+B{r_wcab}/1000)/{f_psat(TCAB)}",
                f"=C{r_wcab}/1000*P_atm/(RATIO_MW+C{r_wcab}/1000)/{f_psat(TCAB)}",
                f"=D{r_wcab}/1000*P_atm/(RATIO_MW+D{r_wcab}/1000)/{f_psat(TCAB)}",
                "—", None,
                "9.1 g/kg at 25 °C is ~46% RH, not the ~40% quoted in doc 20 §5 — still inside the "
                "40–55% comfort band, so no design consequence. See CHK-010", "0.0%")
    r_qwet = mrow("Wet-channel heat Q_wet", f"=B{r_S}*CP_A*({TIN}-B{r_tsup})",
                  f"=C{r_S}*CP_A*({TIN}-C{r_tsup})", f"=D{r_S}*CP_A*({TIN}-D{r_tsup})",
                  "kW", None, "Equation 4 — the heat that must leave evaporatively", "0.00")
    r_hexh = mrow("Exhaust enthalpy (saturated)", f"={f_h(TEXH, 'W_SAT_SINK')}",
                  f"={f_h(TEXH, 'W_SAT_SINK')}", f"={f_h(TEXH, 'W_SAT_SINK')}", "kJ/kg", None,
                  "h_sat at the sink temperature", "0.0")
    r_hent = mrow("Working-air entry enthalpy", f"={f_h(f'B{r_tent}', f'B{r_went}')}",
                  f"={f_h(f'C{r_tent}', f'C{r_went}')}", f"={f_h(f'D{r_tent}', f'D{r_went}')}",
                  "kJ/kg", None, "", "0.0")
    r_dh = mrow("Δh_work", f"=B{r_hexh}-B{r_hent}", f"=C{r_hexh}-C{r_hent}", f"=D{r_hexh}-D{r_hent}",
                "kJ/kg", None, "Equation 5's denominator", "0.0")
    r_M = mrow("Working flow M", f"=B{r_qwet}/B{r_dh}", f"=C{r_qwet}/C{r_dh}", f"=D{r_qwet}/D{r_dh}",
               "kg/s", 0.145, "Published band 0.13–0.16 kg/s (400–500 m³/h)", "0.000")
    r_Mv = mrow("Working flow, volumetric", f"=B{r_M}*SEC_H/RHO_A", f"=C{r_M}*SEC_H/RHO_A",
                f"=D{r_M}*SEC_H/RHO_A", "m³/h", 450, "", "0")
    # duty: process term + working term. X8 recycles the saturated exhaust rather than ambient makeup.
    r_duty = mrow("Sorbent duty (equation 6)",
                  f"=((B{r_S}-B{r_M})*(B{r_wcab}-{WSUP})+B{r_M}*({WAMB}-{WSUP}))*SEC_H/1000",
                  f"=((C{r_S}-C{r_M})*(C{r_wcab}-{WSUP})+C{r_M}*({WAMB}-{WSUP}))*SEC_H/1000",
                  f"=((D{r_S}-D{r_M})*(D{r_wcab}-{WSUP})+D{r_M}*(W_SAT_SINK-{WSUP}))*SEC_H/1000",
                  "kg/h", 10.0, "Published 9–11 kg/h — finding F1", "0.00", 0.15)
    r_regen = mrow("Regeneration heat", f"=B{r_duty}*E_SPEC_S", f"=C{r_duty}*E_SPEC_S",
                   f"=D{r_duty}*E_SPEC_S", "kW", 8.25, "Published 7.5–9 kW", "0.00")
    r_cond = mrow("Moisture desorbed per day", f"=B{r_duty}*24", f"=C{r_duty}*24", f"=D{r_duty}*24",
                  "L/day", None,
                  "Doc 20 §5 publishes 150–220 L/day of RECOVERED condensate — lower than the desorbed "
                  "total because part of the moisture leaves with the ventilation exhaust rather than "
                  "reaching the condenser", "0")
    r_x8 = mrow("X8 closure penalty (vs same-topology open)", "", "",
                f"=D{r_duty}/B{r_duty}", "×", None,
                "Closure re-dries saturated exhaust (25.6 g/kg) instead of taking ambient makeup "
                "(24.2 g/kg), so the working term rises by that ratio alone. Doc 20 §8 quotes a +10–16% "
                "raw swing 'partly bought back by the colder dry-draw supply and the ERV'd ventilation "
                "term' — this sheet lands at about +7%", "0.00")
    r_conv = mrow("Convergence residual (seed − solved cabin ω)", "",
                  f"=C{r_went}-C{r_wcab}", "", "g/kg", None,
                  "Cabin-draw only: how far the seeded working-air entry ω sits from the solved cabin "
                  "state. Re-enter the solved value in the seed cell (or enable iterative calculation) "
                  "until this reads ~0", "0.00")
    put(ws, r_conv, 8, f'=IF(ABS(C{r_conv})<=0.2,"CONVERGED","ITERATE")', "s_calc")
    ws.conditional_formatting.add(f"H{r_conv}", CellIsRule(
        operator="equal", formula=['"CONVERGED"'], fill=PatternFill("solid", fgColor=OKG),
        font=Font(color=OKT, bold=True)))
    ws.conditional_formatting.add(f"H{r_conv}", CellIsRule(
        operator="equal", formula=['"ITERATE"'], fill=PatternFill("solid", fgColor=WARN),
        font=Font(color=WARNT, bold=True)))

    r += 1
    band(ws, r, "READING THE RESULT", 9); r += 1
    put(ws, r, 1,
        "Both open topologies land inside the published 9–11 kg/h band, and the published supply-temperature "
        "band (16.8–18.1 °C) turns out to be exactly the span between them — dry-draw at the cold end, "
        "cabin-draw at the warm end. Cabin-draw is the doctrinal choice (finding X5: route makeup through "
        "the cabin before it becomes working air), and it also delivers the ~600 ppm CO₂ that doc 20 §2 "
        "describes as a structural by-product of open-cycle operation. The X8 column re-dries the saturated "
        "exhaust instead of taking ambient makeup, which costs about +7% on duty here — consistent with "
        "doc 20 §8's '+16% swing, partly bought back'. The reason closure is nearly free at DP-A is visible "
        "in two numbers: ambient sits at 24.2 g/kg and saturated exhaust at 25.6 g/kg. In any milder "
        "ambient that gap widens and closure is what makes the M-cycle water-neutral.", "s_txt")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=9)
    r += 5
    put(ws, r, 1, "Circularity note: equations 1–6 are mutually dependent (S ← T_sup ← ω_work ← M ← Q_wet ← S). "
                  "This sheet closes the loop by fixing each topology's working-air ENTRY state, which is the "
                  "physical design choice, and solving in one pass. Enable iterative calculation in your "
                  "spreadsheet application if you want to relax the cabin-draw entry humidity onto the solved "
                  "cabin state rather than the published one. The full transient model (task T2) supersedes "
                  "this steady state — it does not replace it.", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=9)
    return ws


def sheet_liquid(wb):
    ws = wb.create_sheet("Liquid Sizing")
    ws.sheet_properties.tabColor = "2E75B6"
    titleblock(ws, "Liquid track — mixed-mode sizing (docs 10–12)",
               "The 0.88 kg/h peak, the regeneration budget, the reserve, and the rejection duty, "
               "re-derived live from DP-A. Blue cells are inputs.", 7)
    for col, w in zip("ABCDEFG", (40, 14, 14, 12, 11, 13, 62)):
        ws.column_dimensions[col].width = w
    r = 4
    band(ws, r, "AIR PATH — mixed mode (X6): CO₂-interlocked fresh minimum + recirculated bulk", 7); r += 1
    a = r
    for lbl, val, unit, note in [
        ("Fresh air (CO₂ floor)", "=V_FRESH", "m³/h", "Mechanical minimum stop — never a control setting"),
        ("Fresh air, mass basis", "=B%d*RHO_A_DOC" % r, "kg/h", "Doc 12 quotes ≈58 kg/h"),
        ("Recirculated air", 75, "m³/h", "Doc 10 §2; decouples removal from the ventilation ration"),
        ("Recirculated air, mass basis", "=B%d*RHO_A_DOC" % (r + 2), "kg/h",
         "Doc 12's build-up implies 90.6 kg/h"),
        ("Total absorber flow", "=B%d+B%d" % (r, r + 2), "m³/h", "Doc 12 quotes ~123 m³/h (≈147 kg/h) — see CHK-001"),
        ("Total absorber flow, mass basis", "=B%d+B%d" % (r + 1, r + 3), "kg/h", ""),
    ]:
        kv(ws, r, lbl, val, unit, note, "s_in" if isinstance(val, (int, float)) else "s_calc", "0.0")
        r += 1
    M_FRESH, M_REC, M_TOT = f"B{a+1}", f"B{a+3}", f"B{a+5}"

    r += 1
    band(ws, r, "ABSORBER FLOOR — set by sump concentration, not per-pass approach", 7); r += 1
    b = r
    for lbl, val, unit, note in [
        ("Brine water activity (aw)", "=AW_40", "—", "PENDING test A — ±0.05 aw is ±1.5–3 g/kg of floor"),
        ("Brine temperature", "=T_BRINE", "°C", "Sink + approach; cooling is REQUIRED, not optional (erratum 2)"),
        ("Absorber outlet floor ω", "=1000*RATIO_MW*B%d*%s/(P_atm-B%d*%s)" % (r, f_psat(f'B{r+1}'), r, f_psat(f'B{r+1}')),
         "g/kg", "Doc 12 publishes 11.9 g/kg at aw 0.45 / 30 °C"),
        ("Published floor, base brine", 11.9, "g/kg", "Doc 12 §1 — for comparison"),
        ("Hot-regen floor ω (43–44 wt%)", "=1000*RATIO_MW*AW_44*%s/(P_atm-AW_44*%s)" % (f_psat(f'B{r+1}'), f_psat(f'B{r+1}')),
         "g/kg", "Doc 12 publishes 9.0 g/kg — the deep-dry lever without LiCl"),
    ]:
        style = "s_ref" if lbl.startswith("Published") else ("s_in" if isinstance(val, (int, float)) else "s_calc")
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1
    W_FLOOR = f"B{b+2}"

    r += 1
    band(ws, r, "REMOVAL DUTY — the 0.88 kg/h peak", 7); r += 1
    hdr = r
    for i, t in enumerate(["Term", "Calculated", "Published", "Δ", "Check", "Unit", "Basis"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    delta_row(ws, r, "Fresh-air term, bare intake",
              f"={M_FRESH}*(W_AMB-{W_FLOOR})/1000", 0.713, "kg/h",
              "58 kg/h × (24.2 − 11.9) g/kg — the dominant term without an ERV"); r_f = r; r += 1
    delta_row(ws, r, "Recirculation term",
              f"={M_REC}*(W_CAB_L-{W_FLOOR})/1000", 0.172, "kg/h",
              "90.6 kg/h × (13.8 − 11.9) g/kg — holds the cabin at target"); r_rc = r; r += 1
    delta_row(ws, r, "PEAK REMOVAL, bare fresh",
              f"=B{r_f}+B{r_rc}", 0.88, "kg/h",
              "Doc 12 §1 headline: 0.88 kg/h at DP-A, 4 adults, mixed mode"); r_peak = r; r += 1
    delta_row(ws, r, "Fresh-air term, ERV'd at ε 0.8",
              f"={M_FRESH}*(W_ERV-{W_FLOOR})/1000", 0.232, "kg/h",
              "ω_ERV 15.9 g/kg from doc 12 §1"); r_fe = r; r += 1
    delta_row(ws, r, "PEAK REMOVAL, ERV'd at ε 0.8",
              f"=B{r_fe}+B{r_rc}", 0.49, "kg/h",
              "Doc 12 publishes 0.49 kg/h — which reproduces at ε≈0.65 (ω 17.4), not ε 0.8. See CHK-002",
              tol=0.20); r_peake = r; r += 1
    delta_row(ws, r, "Sanity: flow needed to hold 13.8 g/kg",
              f"=G_LAT/(W_CAB_L-{W_FLOOR})", 147, "kg/h",
              "Gains 280 g/h ÷ (13.8 − 11.9) g/kg — reproduces the 147 kg/h figure"); r += 1

    r += 1
    band(ws, r, "REGENERATION & HEAT REJECTION", 7); r += 1
    hdr = r
    for i, t in enumerate(["Term", "Calculated", "Published", "Δ", "Check", "Unit", "Basis"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    delta_row(ws, r, "Latent duty at peak removal",
              f"=B{r_peak}*H_FG_REGEN/SEC_H", 0.60, "kW", "removal × 2.44 MJ/kg"); r_lat = r; r += 1
    put(ws, r, 1, "COP applied to the peak duty", "s_lbl")
    put(ws, r, 2, 0.65, "s_in", "0.00")
    put(ws, r, 7, "Doc 12 §1 gives a COP build-up of ≈0.59 with no recovery HX and ~0.65 with one. "
                  "The published 0.92 kW headline corresponds to 0.65, i.e. the recovery-HX case — "
                  "see CHK-012", "s_note")
    r_cop = r; r += 1
    delta_row(ws, r, "Regeneration heat, bare fresh",
              f"=B{r_lat}/B{r_cop}", 0.92, "kW",
              "latent ÷ COP; reproduces the published 0.92 kW at COP 0.65"); r_qr = r; r += 1
    delta_row(ws, r, "Regeneration heat, ERV'd",
              f"=B{r_peake}*H_FG_REGEN/SEC_H/B{r_cop}", 0.60, "kW",
              "Does NOT reproduce: the published 0.6 kW needs both the ε 0.65 removal (CHK-002) and a "
              "COP near 0.55. The two ERV'd lines in doc 12 §1 are not on a single consistent basis",
              tol=0.20); r += 1
    delta_row(ws, r, "Absorption heat into the brine",
              f"=B{r_peak}*H_ABS/SEC_H", 0.66, "kW",
              "0.88 kg/h × 2.7 MJ/kg — must be rejected or the contactor self-limits"); r_abs = r; r += 1
    delta_row(ws, r, "Raw-water flow at 2 K rise",
              f"=B{r_abs}/(4.18*2)*SEC_H", 300, "L/h",
              "Doc 12: ~300 L/h base; 3–4 K rise halves the flow at ~1 g/kg of floor"); r += 1
    delta_row(ws, r, "Air-side heat removal, fresh stream only",
              f"={M_FRESH}*CP_A/SEC_H*1000", 16, "W/K",
              "Why an uncooled contactor warms 5–8 K and loses the floor — see CHK-004"); r += 1
    delta_row(ws, r, "Air-side heat removal, total flow",
              f"={M_TOT}*CP_A/SEC_H*1000", 35, "W/K",
              "The ~35 W/K figure quoted in doc 10 §3 is on the mixed-mode flow", tol=0.20); r += 1

    r += 1
    band(ws, r, "MOISTURE BATTERY & DAILY BUDGET", 7); r += 1
    c = r
    for lbl, val, unit, note in [
        ("Water absorbed per kg concentrate", "=C_HI/C_LO-1", "kg/kg", "0.40/0.35 − 1 = 0.143"),
        ("Occupied-night duration", 12, "h", "The reserve must bridge it with zero heat input"),
        ("Duty-averaged fraction of peak", 0.55, "—", "Night duty is below the DP-A peak (sleeping occupants)"),
        ("Overnight concentrate draw", f"=B{r_peak}*B{r+1}*B{r+2}/B{r}", "kg",
         "Doc 10 §3 publishes ~35–40 kg; spec band 25–55 kg"),
        ("Reserve tank spec, minimum", 25, "kg", "Doc 10 §3"),
        ("Reserve tank spec, maximum", 55, "kg", ""),
        ("Duty-schedule factor over 24 h", 0.80, "—",
         "Peak duty does not run all day; the published daily figures are duty-scheduled"),
        ("Daily heat, bare fresh", f"=B{r_qr}*24*B{r+6}", "kWh/day", "Doc 12 publishes 17–20 kWh/day"),
        ("ERV + DCV reduction factor", 0.54, "—",
         "Ratio of the published ERV'd budget to the bare-fresh budget (10 ÷ 18.5 at midpoints)"),
        ("Daily heat, ERV + DCV scheduled", "", "kWh/day",
         "Doc 12 publishes 9–11 kWh/day — the budget the solar array is sized to"),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1
    put(ws, c + 9, 2, f"=B{c+7}*B{c+8}", "s_calc", "0.00")
    put(ws, c + 9, 4, "Duty-scheduling and DCV take the bare-fresh budget down to the published 9–11 kWh/day; "
                      "the ERV is ~9–10 kWh/day of that saving on its own (X7)", "s_note")

    r += 1
    band(ws, r, "CONTACTOR SIZING", 7); r += 1
    for lbl, val, unit, note in [
        ("Cell rated capacity (mid-band)", 0.6, "kg/h", "PENDING test I — published band 0.4–0.8 kg/h"),
        ("Cells required at peak duty", f"=CEILING(B{r_peak}/B{r},1)", "cells",
         "Doc 11: mixed-mode duty sits inside a 2-cell bank at mid-rating"),
        ("Bank provision (+1 empty bay)", f"=B{r+1}+1", "bays", "One deliberately empty bay is design doctrine"),
        ("Irrigation per cell", 9, "L/min", "150–240 L/min·m²; underwetting collapses K·a nonlinearly"),
        ("Total irrigation demand", f"=B{r+1}*B{r+3}", "L/min", "The binding constraint is wetting, not airflow"),
        ("Face velocity", 0.8, "m/s", "Band 0.6–1.0 m/s"),
        ("Face area implied by total flow", f"={M_TOT}/RHO_A/SEC_H/B{r+5}", "m²", ""),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1
    return ws


def sheet_solid(wb):
    ws = wb.create_sheet("Solid Sizing")
    ws.sheet_properties.tabColor = "548235"
    titleblock(ws, "Solid track — DCHX sizing (docs 20–22)",
               "Duty → sorbent inventory → coated area → condensate balance, with the F3 effective-Δq "
               "penalty applied explicitly rather than buried.", 7)
    for col, w in zip("ABCDEFG", (40, 14, 14, 12, 11, 13, 62)):
        ws.column_dimensions[col].width = w
    r = 4
    band(ws, r, "DUTY (from the Model 00§4 sheet — cabin-draw solution)", 7); r += 1
    d = r
    for lbl, val, unit, note in [
        ("Peak sorbent duty", 10.0, "kg/h", "Published band 9–11 kg/h (F1); the model sheet re-derives it"),
        ("Half-cycle time", 10, "min", "Throughput is governed by cycle time and coated area, not mass"),
        ("Water per half-cycle", f"=B{r}*B{r+1}/60", "kg", ""),
        ("Modules alternating", 2, "—", "One adsorbing while the other regenerates"),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1
    DUTY, W_CYC = f"B{d}", f"B{d+2}"

    r += 1
    band(ws, r, "SORBENT INVENTORY — the F3 penalty made explicit", 7); r += 1
    hdr = r
    for i, t in enumerate(["Basis", "Δq (g/g)", "Inventory (kg)", "Coated area (m²)",
                           "vs envelope-only", "Grade", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    load = 0.18  # kg sorbent per m² planform
    kv_rows = [
        ("Envelope-only (cautionary row)", 0.25, 2.5, "superseded",
         "The naive sizing error: ignores where the sensible heat goes"),
        ("Self-consistent peak, dry-purge Δq", 0.25, None, "sizing-grade",
         "Published 0.2–0.3 g/g — a DRY-PURGE full-swing figure (F3)"),
        ("Self-consistent peak, effective Δq", 0.175, None, "PENDING M2",
         "0.15–0.2 g/g after a 0.05–0.10 g/g residual at 60–65 °C — the real sizing number"),
        ("Conservative (Δq floor)", 0.15, None, "PENDING M2", "The bottom of the F3 band"),
    ]
    first = r
    for lbl, dq, duty_override, grade, note in kv_rows:
        put(ws, r, 1, lbl, "s_lbl")
        put(ws, r, 2, dq, "s_in", "0.000")
        duty_ref = duty_override if duty_override is not None else None
        if duty_ref is not None:
            put(ws, r, 3, f"={duty_ref}*B{d+1}/60/B{r}*B{d+3}", "s_calc", "0.0")
        else:
            put(ws, r, 3, f"={W_CYC}/B{r}*B{d+3}", "s_calc", "0.0")
        put(ws, r, 4, f"=C{r}/{load}", "s_calc", "0.0")
        put(ws, r, 5, f"=IF($C${first}=0,\"\",C{r}/$C${first})", "s_calc", "0.0\"×\"")
        put(ws, r, 6, grade, "s_txt")
        put(ws, r, 7, note, "s_note")
        r += 1
    put(ws, r, 1, f"Doc 22 §4 states the self-consistent charge is 3–5× the envelope-only row (~3–4 kg / 10–13 m²), "
                  f"then +25–50% again once F3 compounds. The envelope-only inventory reproduces at "
                  f"{load} kg/m² planform (doc 22 §2), but the published 10–13 m² for that 3–4 kg implies "
                  f"~0.3 kg/m² — the coated-area column therefore does not reconcile with the stated "
                  f"loading. See CHK-013.", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 2

    band(ws, r, "ENERGY & WATER BALANCE", 7); r += 1
    hdr = r
    for i, t in enumerate(["Term", "Calculated", "Published", "Δ", "Check", "Unit", "Basis"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    delta_row(ws, r, "Latent floor of desorption", f"={DUTY}*0.67", 6.7, "kW",
              "0.67 kWh/L — material-independent, 70–85% of the input"); r += 1
    delta_row(ws, r, "Regeneration heat, total", f"={DUTY}*E_SPEC_S", 8.25, "kW",
              "0.85 kWh/L × duty; published band 7.5–9 kW"); r_reg = r; r += 1
    delta_row(ws, r, "Daily regeneration heat", f"=B{r_reg}*24", 198, "kWh/day",
              "Continuous duty at DP-A — why this track is waste-heat-coupled"); r += 1
    delta_row(ws, r, "Moisture desorbed per day", f"={DUTY}*24", None, "L/day",
              "Doc 20 §5 publishes 150–220 L/day of RECOVERED condensate. The desorbed total is higher "
              "because part of the moisture leaves with the ventilation exhaust rather than reaching "
              "the condenser train"); r_c = r; r += 1
    delta_row(ws, r, "Condensate recovered (published basis)", 185, None, "L/day",
              "Midpoint of the published 150–220 L/day"); r_cr = r; r += 1
    delta_row(ws, r, "Potable-grade surplus", 60, None, "L/day",
              "Published 50–70 L/day: the gains + ventilation terms"); r_su = r; r += 1
    delta_row(ws, r, "M-cycle wet-channel feed", f"=B{r_cr}-B{r_su}", None, "L/day",
              "What the condensate must cover for the loop to close"); r += 1
    delta_row(ws, r, "DHW heat recoverable", 2.0, None, "kWh/day",
              "Demand-capped 1–3 kWh/day — first in the doc 00 §7 rule-4 recapture order"); r_dhw = r; r += 1
    delta_row(ws, r, "DHW recovery as a fraction of input", f"=B{r_dhw}/(B{r_reg}*24)", None, "—",
              "Doc 20 §3 quotes '~10–15% recoverable as useful energy' but the DHW cap it cites in the "
              "same sentence (1–3 kWh/day) is only ~1% of the daily regeneration input at this duty. "
              "The two statements are on different bases — see CHK-015"); r += 1

    r += 1
    band(ws, r, "REGENERATION AGAINST HUMID PURGE — why 60–65 °C, not 50 °C (F2)", 7); r += 1
    hdr = r
    for i, t in enumerate(["Bed temperature (°C)", "P_sat at bed (kPa)", "Purge P_v (kPa)",
                           "RH at bed face", "Doc 20 §6", "vs AlFu step ~25–30%", "Verdict"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    doc_face = {45: 0.38, 50: 0.29, 60: 0.17, 65: 0.13}
    for t in (45, 50, 55, 60, 65, 70):
        put(ws, r, 1, t, "s_in", "0")
        put(ws, r, 2, f"={f_psat(f'A{r}')}", "s_calc", "0.00")
        put(ws, r, 3, f"={f_pv('SD_W_PURGE/1000')}", "s_calc", "0.000")
        put(ws, r, 4, f"=C{r}/B{r}", "s_calc", "0.0%")
        put(ws, r, 5, doc_face.get(t, ""), "s_ref", "0%")
        put(ws, r, 6, f'=IF(D{r}>=0.30,"above the step","" ) & IF(AND(D{r}<0.30,D{r}>=0.25),"on the step","") & IF(D{r}<0.25,"below the step","")', "s_calc")
        put(ws, r, 7, f'=IF(D{r}>=0.30,"zero driving force",IF(D{r}>=0.25,"marginal","works"))', "s_calc")
        ws.conditional_formatting.add(f"G{r}", CellIsRule(
            operator="equal", formula=['"zero driving force"'],
            fill=PatternFill("solid", fgColor=BADF), font=Font(color=BADT, bold=True)))
        ws.conditional_formatting.add(f"G{r}", CellIsRule(
            operator="equal", formula=['"works"'],
            fill=PatternFill("solid", fgColor=OKG), font=Font(color=OKT, bold=True)))
        ws.conditional_formatting.add(f"G{r}", CellIsRule(
            operator="equal", formula=['"marginal"'],
            fill=PatternFill("solid", fgColor=WARN), font=Font(color=WARNT, bold=True)))
        r += 1
    put(ws, r, 1, "Purge humidity ω (condensing purge)", "s_lbl")
    put(ws, r, 2, 25, "s_in", "0.0")
    put(ws, r, 3, "g/kg")
    put(ws, r, 4, "The single input that decides the whole heat-source question. AlFu's often-quoted "
                  "'~50 °C regeneration' is a DRY-purge figure and does not transfer. M2 is run against "
                  "a logged ~24 g/kg purge, never dry lab air. Note: doc 20 §6's own RH column "
                  "back-solves to ~23 g/kg rather than the ~25 g/kg its text states — at 23 g/kg the "
                  "50 °C row reads 29% and is 'on the step' rather than above it (CHK-012).", "s_note")
    ws.merge_cells(start_row=r, start_column=4, end_row=r + 1, end_column=7)
    wb.defined_names["SD_W_PURGE"] = DefinedName("SD_W_PURGE", attr_text=f"'Solid Sizing'!$B${r}")
    r += 3

    band(ws, r, "SYNTHESIS SCALING (doc 21 §3, aqueous route) — 1 Al : 1 fumarate : 2 NaOH", 7); r += 1
    hdr = r
    for i, t in enumerate(["Reagent", "MW (g/mol)", "Validation ~19 g", "Bench ~75 g",
                           "Scaled to charge", "Unit", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    target = r + 5
    for name, mw, val, bench, note in [
        ("Al₂(SO₄)₃·18H₂O", 666.43, 41.7, 166.6, "Verify hydrate assay per lot COA"),
        ("Fumaric acid", 116.07, 14.5, 58.0, "FCC / E297 grade"),
        ("NaOH", 40.00, 10.0, 40.0, "Add to water slowly — strongly exothermic"),
        ("DI water", None, 190, 750, "mL, not g"),
    ]:
        put(ws, r, 1, name, "s_lbl")
        put(ws, r, 2, mw, "s_num", "0.00")
        put(ws, r, 3, val, "s_ref", "0.0")
        put(ws, r, 4, bench, "s_ref", "0.0")
        put(ws, r, 5, f"=D{r}*$B${target}/0.075", "s_calc", "0.0")
        put(ws, r, 6, "g" if mw else "mL")
        put(ws, r, 7, note, "s_note")
        r += 1
    put(ws, r, 1, "Mole check (bench): 0.50 mol Al : 0.50 mol fumarate : 1.00 mol NaOH = 1 : 1 : 2", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 1
    kv(ws, r, "Target MOF mass", 2.5, "kg", "Full charge is ~2.5 kg indicative; hand-grinding does not scale to it")
    return ws


def sheet_co2(wb):
    ws = wb.create_sheet("CO2 & Ventilation")
    ws.sheet_properties.tabColor = "BF8F00"
    titleblock(ws, "CO₂ & ventilation — the four-layer stack (doc 00 §5, spec P17)",
               "SAFETY-CRITICAL. <1,000 ppm at all times in every mode including sealed; "
               "alarm and forced boost at 2,000 ppm; per-room maximum sensing; mechanical minimum stop.", 8)
    for col, w in zip("ABCDEFGH", (36, 14, 14, 12, 11, 12, 13, 58)):
        ws.column_dimensions[col].width = w
    r = 4
    band(ws, r, "DOSE–RESPONSE LADDER — c = c_out + 10⁶·V̇_CO₂/V̇_fresh", 8); r += 1
    hdr = r
    for i, t in enumerate(["Fresh air (m³/h)", "Per person (m³/h)", "CO₂ awake (ppm)",
                           "Doc 00 §5", "Δ", "CO₂ asleep (ppm)", "vs spec", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    doc_ppm = {48: 1920, 36: 2420, 24: 3420, 12: 6420}
    for v in (124, 91, 48, 36, 24, 12, 6):
        put(ws, r, 1, v, "s_in", "0")
        put(ws, r, 2, f"=A{r}/N_OCC", "s_calc", "0.0")
        put(ws, r, 3, f"=CO2_OUT+1000000*CO2_GEN/A{r}", "s_calc", "0")
        put(ws, r, 4, doc_ppm.get(v, ""), "s_ref", "0")
        put(ws, r, 5, f'=IF(D{r}="","",(C{r}-D{r})/D{r})', "s_calc", "0.0%")
        put(ws, r, 6, f"=CO2_OUT+1000000*0.047/A{r}", "s_calc", "0")
        put(ws, r, 7, f'=IF(C{r}<=CO2_SPEC,"PASS",IF(C{r}>=2000,"ALARM","OVER SPEC"))', "s_calc")
        note = {48: "The ventilation floor — above spec on its own; the layer stack closes it",
                124: "X8 mode: ventilation sized for the spec, +10–15% duty at solid-track scale",
                12: "Per-person floor if the interlock were defeated",
                6: "Illustrative only — below any legal minimum"}.get(v, "")
        put(ws, r, 8, note, "s_note")
        for txt, fill, fg in (("PASS", OKG, OKT), ("OVER SPEC", WARN, WARNT), ("ALARM", BADF, BADT)):
            ws.conditional_formatting.add(f"G{r}", CellIsRule(
                operator="equal", formula=[f'"{txt}"'],
                fill=PatternFill("solid", fgColor=fill), font=Font(color=fg, bold=True)))
        r += 1
    put(ws, r, 1, "Sealed envelope (infiltration only) reads ~7,600 ppm; a closed room runs "
                  "+2,000–2,900 ppm over the main space, which is why sensing is per-room and the "
                  "interlock keys to the MAXIMUM reading, not an average.", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 2

    band(ws, r, "LAYER 1 — ERV ECONOMICS (X7)", 8); r += 1
    e = r
    for lbl, val, unit, note in [
        ("ERV latent effectiveness ε", "=EPS_LAT", "—", "PENDING test E; purchasable with membrane area"),
        ("Moisture gradient ambient − cabin", "=W_AMB-W_CAB_L", "g/kg", "Doc 00 quotes 10.4 g/kg"),
        ("Unrecovered moisture per kg fresh", f"=(1-B{r})*B{r+1}", "g/kg", "The (1−ε)×10.4 term"),
        ("Marginal heat, ERV'd", f"=B{r+2}/1000*H_FG_REGEN/3600*24/COP_REGEN*RHO_A", "kWh/day per m³/h",
         "Doc 00 quotes 0.06 — this sheet reproduces the order"),
        ("Marginal heat, bare fresh", f"=B{r+1}/1000*H_FG_REGEN/3600*24/COP_REGEN*RHO_A", "kWh/day per m³/h",
         "Doc 00's '3.8 kWh/day per 12 m³/h' is this bare figure — see CHK-003"),
        ("Heat saved per 12 m³/h, bare basis", f"=B{r+4}*12", "kWh/day", "Doc 00 quotes ~3.8 kWh/day"),
        ("ERV saving at the 48 m³/h floor", f"=(B{r+4}-B{r+3})*V_FRESH", "kWh/day",
         "Doc 11 §5 / X7 quotes ~9–10 kWh/day. This sheet lands higher because it runs the full "
         "peak-duty gradient for 24 h; the published figure is duty-scheduled, like the daily budgets "
         "on the Liquid Sizing sheet. Either way it is the single largest lever on the heat budget"),
        ("CO₂ crossover (EATR) limit", 0.05, "—", "Acceptance limit, test E — an ERV must not recycle CO₂"),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.000")
        r += 1

    r += 1
    band(ws, r, "LAYER 2 — THE CO₂ BATTERY (X10 amine / X11 potash) — REQUIRED-PENDING tests J / J-K", 8); r += 1
    hdr = r
    for i, t in enumerate(["Quantity", "Amine (solar grade)", "Potash (waste-heat)", "Unit",
                           "Published", "Check", "", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    for lbl, am, po, unit, pub, note in [
        ("Regeneration temperature", "85–95", "130–150", "°C", "",
         "Potash is equilibrium-dead below ~120 °C; the ETC's 93 °C ceiling cannot regenerate it"),
        ("Scrubbing duty at the floor", "=1.6", "=1.6", "kg/day", 1.6,
         "Chemistry-agnostic: the duty is set by the crew and the ventilation rate"),
        ("Bed mass, each of two", "=3", "=4.5", "kg", "", "Amine ~3 kg; potash 4–5 kg at 0.5–0.65 mmol/g working"),
        ("Half-cycle", "=90", "=90", "min", 90, ""),
        ("Recirculation through the bed", "=91", "=91", "m³/h", 91, "At 50% single-pass capture; +50–150 Pa"),
        ("Specific regeneration energy", "=1.15", "", "kWh/kg", "", "Amine 1.0–1.3 incl. water co-adsorption + sensible"),
        ("Daily regeneration heat", "=B%d*B%d" % (r + 1, r + 5), "", "kWh/day", 2.05,
         "Published 1.6–2.5 kWh/day — schedulable into the solar window"),
        ("Overnight stored obligation", "=0.3", "=0.3", "kg", 0.3, "Loaded beds carry the night"),
        ("Bed placement humidity", "45–55", "45–55", "% RH", "",
         "Post-dehumidification. Carbonation WANTS this band — water is a reagent"),
        ("Breathing-air gate", "amine/ammonia slip (J)", "alkaline carryover (J-K)", "—", "",
         "Safety register item 4 — gates any connection to breathing air"),
        ("Support prohibition", "—", "alumina, MgO", "—", "", "Irreversible double-salt deactivation"),
    ]:
        put(ws, r, 1, lbl, "s_lbl")
        put(ws, r, 2, am, "s_calc" if isinstance(am, str) and am.startswith("=") else "s_txt", "0.00")
        put(ws, r, 3, po, "s_calc" if isinstance(po, str) and po.startswith("=") else "s_txt", "0.00")
        put(ws, r, 4, unit)
        put(ws, r, 5, pub, "s_ref", "0.00")
        put(ws, r, 8, note, "s_note")
        r += 1

    r += 1
    band(ws, r, "LAYER 0 — X9: WHY THERE IS NO GAS IN THE ENVELOPE", 8); r += 1
    for lbl, val, unit, note in [
        ("Crew CO₂, 4 occupants awake", "=CO2_GEN", "m³/h", "The whole design load"),
        ("One 2 kW gas burner", 0.26, "m³/h", "3.6× the entire crew"),
        ("Burner ratio to crew", f"=B{r+1}/B{r}", "×", "Doc 00 §5 quotes 3.6×"),
        ("Burner combustion water", 0.25, "kg/h", "A latent load on top of the CO₂"),
        ("Induction galley cost", "=2", "kWh_e/day", "Deletes both terms — safety register item 2"),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1
    r += 1
    put(ws, r, 1, "LAYER 3 — INTERLOCK: CO₂-governed fresh damper with a MECHANICAL minimum stop, so no "
                  "controller fault or economy setting can close ventilation below the floor. "
                  "Recirculation-only operation is prohibited while occupied. Scrubbing never substitutes "
                  "for the ventilation floor. Each 12 m³/h of fresh air cut saves ~3.8 kWh/day of heat — a "
                  "standing economic temptation that the mechanical stop, not software, forecloses.", "s_txt")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=8)
    return ws


def sheet_heatwater(wb):
    ws = wb.create_sheet("Heat & Water")
    ws.sheet_properties.tabColor = "C55A11"
    titleblock(ws, "Platform heat cascade & water ladder (doc 30)",
               "Match each load to the grade of heat it needs; spend high grade only at the cascade bottom. "
               "Water comes from heat, air, or sky — not meaningfully from electricity.", 7)
    for col, w in zip("ABCDEFG", (38, 14, 14, 14, 12, 16, 58)):
        ws.column_dimensions[col].width = w
    r = 4
    band(ws, r, "HEAT-GRADE CASCADE", 7); r += 1
    hdr = r
    for i, t in enumerate(["Rung", "Grade (°C)", "Load", "Served by", "Note", "", ""], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    for rung, grade, load, src, note in [
        ("Source", "high (continuous)", "Electricity generation", "High-grade process heat",
         "Single point of failure for water AND power simultaneously"),
        ("1", "130–150", "K₂CO₃ CO₂-bed regeneration (X11)", "Process tail",
         "Only on platforms with a ≥~130 °C tap"),
        ("2", "75–85", "Two-stage HDH humidifier", "Process tail",
         "The water engine at saline-feed sites; never pointed at cabin air"),
        ("3", "85–95", "Amine CO₂ bed + hot-regen still", "Process tail · ETC · X12 AHT",
         "The grade an AHT can synthesise from a 60–65 °C tail"),
        ("4", "60–93", "Liquid-track regeneration", "ETC solar 2.5–4.5 m²",
         "The solar comfort island — X2 keeps driving force positive across the whole band"),
        ("5", "60–65", "Solid-bed regeneration", "Waste-heat tail (primary)",
         "F2: solar-direct is equilibrium-dead here against a condensing purge"),
        ("6", "45–50", "PVT thermal output", "PVT ×2",
         "Cannot reach HDH grade; F2-dead for direct solid-bed regeneration"),
        ("Sink", "~29", "Rejection", "Raw water",
         "Adsorption heat releases only a few K above the sink — reject it, don't chase it"),
    ]:
        put(ws, r, 1, rung, "s_lbl"); put(ws, r, 2, grade); put(ws, r, 3, load)
        put(ws, r, 4, src); put(ws, r, 5, note, "s_note")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        r += 1

    r += 1
    band(ws, r, "LIQUID-TRACK DAILY HEAT BUDGET & SOURCE SIZING", 7); r += 1
    h = r
    for lbl, val, unit, note in [
        ("Daily regeneration heat (ERV + DCV)", 10, "kWh/day", "Published band 9–11 kWh/day"),
        ("CO₂-battery regeneration", 2.05, "kWh/day", "Published 1.6–2.5 kWh/day (test J)"),
        ("Domestic hot water draw", 2, "kWh/day", "Demand-capped 1–3 kWh/day — first in the recapture order"),
        ("Total thermal demand", f"=SUM(B{r}:B{r+2})", "kWh/day", ""),
        ("Solar yield per m² per day", 2.75, "kWh/m²·day", "2.5–3 kWh/m²·day basis"),
        ("Collector area required", f"=B{r+3}/B{r+4}", "m²", "Published 2.5–4.5 m² flat-plate class"),
        ("Diesel heating value", 8.5, "kWh/L", "5 kW hydronic heater class"),
        ("Diesel-only fuel burn", f"=B{r+3}/B{r+6}", "L/day", "Published 1.1–1.4 L/day at full DP-A duty"),
        ("Induction galley electrical", 2, "kWh_e/day", "X9 — a battery/inverter sizing line on a solar-only island"),
        ("System electrical, scaled", 65, "W", "Published 50–80 W band including ERV and exhaust fans"),
        ("Daily electrical, comfort system", f"=B{r+9}*24/1000", "kWh_e/day", ""),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1

    r += 1
    band(ws, r, "WATER REDUNDANCY LADDER — independence is the property that matters", 7); r += 1
    hdr = r
    for i, t in enumerate(["Path", "Source class", "Independent of HDH?", "Rate", "Unit", "Role", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    for path, cls, indep, rate, unit, role, note in [
        ("Process-heat HDH", "raw water + free heat", "— (is the HDH)", "high", "L/day", "primary",
         "GOR 2.2–2.8 at a 29 °C sink; ~90 L/day per kW of heat"),
        ("Resistive-HDH", "raw water + battery electricity", "NO — common mode", "high", "L/day",
         "heat-source backup only", "Shares columns, raw-water loop, and air loop; ~27× RO's energy per litre"),
        ("Solid-track condensate", "humid air + low-grade heat", "YES", "50–70", "L/day",
         "comfort byproduct", "X8 custody; water-neutral M-cycle in every ambient"),
        ("Liquid-track distillate", "humid air + solar-grade heat", "YES", "8–18", "L/day",
         "solar-independent path", "Also the M-cycle feed — the loop closes"),
        ("Rain catchment", "sky", "YES", "generous, intermittent", "—", "passive",
         "Over-invest here: it is the cheapest independent path"),
        ("Emergency RO (pickled)", "raw water + small power", "YES", "drinking + essentials", "—",
         "dormant", "The only path robust OUTSIDE the humid tropics; few cycles → minimal fouling"),
        ("Fresh-water tanks", "—", "YES", "buffer", "—", "autonomy",
         "Bridges any outage; the other cheap independent path"),
    ]:
        put(ws, r, 1, path, "s_lbl"); put(ws, r, 2, cls); put(ws, r, 3, indep)
        put(ws, r, 4, rate); put(ws, r, 5, unit); put(ws, r, 6, role); put(ws, r, 7, note, "s_note")
        ws.conditional_formatting.add(f"C{r}", CellIsRule(
            operator="containsText", formula=[f'NOT(ISERROR(SEARCH("NO",C{r})))'],
            fill=PatternFill("solid", fgColor=BADF), font=Font(color=BADT, bold=True)))
        r += 1
    put(ws, r, 1, "Common-mode warning: process-heat HDH and resistive-HDH share columns, the raw-water loop, and "
                  "the air loop — resistive backs up heat-source loss only. Real security lives in the "
                  "mechanically independent paths, and the cheapest of those (tankage and rain) is where to "
                  "over-invest. Redundancy outranks efficiency in the field.", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=7)
    return ws


def sheet_upgrades(wb):
    ws = wb.create_sheet("Upgrade Paths")
    ws.sheet_properties.tabColor = "7F7F7F"
    titleblock(ws, "Upgrade paths — boost modes only (doc 31)",
               "BINDING RULE: every path here is an addition, never core. No DP-A comfort, air-quality, or "
               "water claim may depend on any of them; the baseline must remain fully operable with every "
               "one removed. All figures estimate-grade — nothing here has been bench-tested.", 8)
    for col, w in zip("ABCDEFGH", (36, 14, 14, 14, 12, 13, 13, 56)):
        ws.column_dimensions[col].width = w
    r = 4
    band(ws, r, "X12 — CaCl₂ ABSORPTION HEAT TRANSFORMER · the lift-ceiling inequality, solved live", 8); r += 1
    put(ws, r, 1, "Governing inequality:  aw(x) · P_sat(T_abs) < P_sat(T_evap)   — absorption stays "
                  "downhill only while the brine's depressed vapour pressure sits below the evaporator's.", "s_txt")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 2
    hdr = r
    for i, t in enumerate(["Strong brine", "aw", "T_evap 60 °C → T_abs max",
                           "Doc 31 §2.3", "T_evap 65 °C → T_abs max", "Doc 31 §2.3",
                           "Check (65 °C)", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    for label, aw, d60, d65, note in [
        ("44 wt%", 0.34, 85.5, 91, "The deliverable: a ~20–25 K gross lift into the 85–90 °C band"),
        ("43 wt%", 0.37, 83, 88.5, ""),
        ("42 wt%", 0.40, 81, 86.5, "Below hot-regen still grade — concentration is the lever"),
        ("40 wt%", 0.45, 77, 83, "Base brine cannot reach amine-bed grade"),
    ]:
        put(ws, r, 1, label, "s_lbl")
        put(ws, r, 2, aw, "s_in", "0.00")
        put(ws, r, 3, f"={f_tsat_at(f'{f_psat(60)}/B{r}')}", "s_calc", "0.0")
        put(ws, r, 4, d60, "s_ref", "0.0")
        put(ws, r, 5, f"={f_tsat_at(f'{f_psat(65)}/B{r}')}", "s_calc", "0.0")
        put(ws, r, 6, d65, "s_ref", "0.0")
        put(ws, r, 7, f'=IF(ABS((E{r}-F{r})/F{r})<=0.03,"OK","REVIEW")', "s_calc")
        put(ws, r, 8, note, "s_note")
        ws.conditional_formatting.add(f"G{r}", CellIsRule(
            operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=OKG),
            font=Font(color=OKT, bold=True)))
        ws.conditional_formatting.add(f"G{r}", CellIsRule(
            operator="equal", formula=['"REVIEW"'], fill=PatternFill("solid", fgColor=WARN),
            font=Font(color=WARNT, bold=True)))
        r += 1
    put(ws, r, 1, "The ceiling is real and single-stage: the X11 potash rung (130–150 °C) is beyond reach. "
                  "Test A's measured aw table feeds this inequality directly — the ±0.05 aw uncertainty on "
                  "the brine side propagates straight into the lift ceiling.", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 2

    band(ws, r, "X12 — ENERGY BALANCE", 8); r += 1
    u = r
    for lbl, val, unit, note in [
        ("AHT thermal COP", 0.465, "—", "0.45–0.48 before losses: upgraded heat ÷ driving heat"),
        ("Upgraded heat required (amine bed)", 2.25, "kWh/day", "The CO₂ battery's 1.6–2.5 kWh/day at 85–95 °C"),
        ("Driving tail heat drawn", f"=B{r+1}/B{r}", "kWh/day", "Doc 31 quotes ~5 kWh/day"),
        ("Heat rejected via the still condenser", f"=B{r+2}-B{r+1}", "kWh/day", "Through hardware already running"),
        ("Continuous form", f"=1/B{r}", "kW tail per kW upgraded", "Doc 31 quotes ~2.2"),
        ("Distillate lift pump head", 16, "kPa", "Watt-scale peristaltic — unlike LiBr's solution pump"),
        ("Absorption latent release", 2.6, "MJ/kg", "Released AT the brine's temperature — the X12 finding"),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1

    r += 1
    band(ws, r, "STILL MVR — the pressure ratio decomposition (doc 31 §4)", 8); r += 1
    m = r
    for lbl, val, unit, note in [
        ("Pool temperature", 70, "°C", ""),
        ("Brine concentration", 42, "wt%", "aw ≈ 0.40 at this concentration"),
        ("Water activity at that concentration", 0.40, "—", "PENDING test A"),
        ("Pure-water saturation at pool T", f"={f_psat(f'B{m}')}", "kPa", "What the headspace would be over pure water"),
        ("Actual headspace pressure", f"=B{m+3}*B{m+2}", "kPa", "Doc 31 quotes ~12.5 kPa"),
        ("Required in-pool condensing pressure", 40, "kPa", "Sat. 76–78 °C for a useful 6–8 K ΔT"),
        ("Pressure ratio", f"=B{m+5}/B{m+4}", "—", "Doc 31 quotes ~3"),
        ("Ratio attributable to activity depression", f"=B{m+3}/B{m+4}", "—",
         "The machine pays for the desiccant's aw as pressure ratio — the compression-side mirror of X12"),
        ("Specific compression work", 140, "Wh/kg", "vs ~1,800 Wh/kg thermal"),
        ("Electrical at 0.88 kg/h peak duty", f"=B{m+8}*0.88", "W", "Doc 31 quotes 110–140 W — PV-scale"),
    ]:
        style = "s_in" if isinstance(val, (int, float)) else "s_calc"
        kv(ws, r, lbl, val, unit, note, style, "0.00")
        r += 1

    r += 1
    band(ws, r, "THE FAMILY AT A GLANCE — and the recorded rejection", 8); r += 1
    hdr = r
    for i, t in enumerate(["Path", "Character", "Converts", "Into", "Cost", "First gate", "Platform trigger", "Note"], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    for row in [
        ("X12 AHT (§2)", "heat-only, no new chemistry", "60–65 °C waste tail", "85–90 °C process heat",
         "COP 0.45–0.48", "test L", "waste-heat tail ≤65 °C", "PRIMARY upgrade — the still is half the machine"),
        ("Coupled VC heat pump (§3)", "electron-rich boost", "~350–500 W electric",
         "brine floor depth + regen heat", "+3.5–4.5 g/kg", "procurement + test I",
         "electron-rich, heat-poor", "Most of the LiCl benefit for ~400 W and no new chemistry"),
        ("Still MVR (§4)", "heat-outage regenerator", "~110–140 W electric", "full still duty, no heat source",
         "120–160 Wh/kg", "compressor mist pre-gate", "electron-rich", "Makes the moisture battery rechargeable on electricity alone"),
        ("Closed AlFu chiller (§5)", "heat-rich boost", "2–2.5 kW solar/waste heat", "heat-driven brine chilling",
         "COP_th 0.4–0.5", "solid M-series + leak-down", "heat-rich, electricity-poor",
         "Closed cycle → no purge → X2-clean: the solid chemistry finally works on solar heat"),
        ("Static crystallizer (§6)", "mass-limited platforms", "reserve tank mass", "3–5× denser moisture battery",
         "35–40 → 10–15 kg", "test A3 (~$10)", "marine mass-limited", "Crystals never travel — the pot is valve-isolated"),
        ("REJECTED: heat pump as prime mover", "—", "~2 kW electric", "regeneration-grade heat", "COP_h 3.5–4",
         "—", "none", "Loses outright to a 1–3 kW VC-AC and abandons the heat-flexibility proposition"),
    ]:
        for i, v in enumerate(row, start=1):
            put(ws, r, i, v, "s_lbl" if i == 1 else ("s_note" if i == 8 else "s_txt"))
        if row[0].startswith("REJECTED"):
            for i in range(1, 9):
                ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GREY)
        r += 1
    return ws


def sheet_tests(wb):
    ws = wb.create_sheet("Test Program")
    ws.sheet_properties.tabColor = "375623"
    titleblock(ws, "Test program — cheapest decisive experiment first",
               "Every PENDING grade in the register points at a row here. The budget sums live: "
               "change the in-baseline flag and the total follows.", 9)
    cols = [("ID", 8), ("Track", 10), ("Title", 40), ("Cost lo", 10), ("Cost hi", 10),
            ("Duration", 12), ("In baseline", 12), ("What it measures", 62), ("What it decides", 62)]
    header(ws, cols, row=5)
    r = 6
    first = r
    for tid, track, title, lo, hi, dur, base, meas, dec in TESTS:
        put(ws, r, 1, tid, "s_lbl"); put(ws, r, 2, track); put(ws, r, 3, title)
        put(ws, r, 4, lo, "s_num", "$#,##0"); put(ws, r, 5, hi, "s_num", "$#,##0")
        put(ws, r, 6, dur); put(ws, r, 7, "yes" if base else "no", "s_num")
        put(ws, r, 8, meas); put(ws, r, 9, dec)
        if base:
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor=OKG)
        r += 1
    last = r - 1
    ws.auto_filter.ref = f"A5:I{last}"
    ws.freeze_panes = "C6"
    # live budget block
    put(ws, 3, 1, "Baseline gating budget", "s_lbl")
    put(ws, 3, 4, f'=SUMIF($G${first}:$G${last},"yes",D{first}:D{last})', "s_calc", "$#,##0")
    put(ws, 3, 5, f'=SUMIF($G${first}:$G${last},"yes",E{first}:E{last})', "s_calc", "$#,##0")
    put(ws, 3, 6, "USD", "s_txt")
    put(ws, 3, 8, "Live sum of the tests flagged in-baseline above. The former '~$485–755' "
                  "aggregate headline was retired in doc 12 v1.4 — it did not decompose from the "
                  "table and a currency-and-date-specific total dates badly in a prior-art record. "
                  "Read every cost here as 2026 order-of-magnitude, one currency, one region.", "s_note")
    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=9)
    put(ws, 4, 1, "All listed tests", "s_lbl")
    put(ws, 4, 4, f"=SUM(D{first}:D{last})", "s_calc", "$#,##0")
    put(ws, 4, 5, f"=SUM(E{first}:E{last})", "s_calc", "$#,##0")
    put(ws, 4, 6, "USD", "s_txt")
    put(ws, 4, 8, "Including the platform-conditional entries (A2, A3, C, J-K, L) that join the set "
                  "only when their trigger exists. Solid-track T/M entries carry no cost line in the "
                  "documents — they are effort, plus outsourced PXRD/DVS fees.", "s_note")
    ws.merge_cells(start_row=4, start_column=8, end_row=4, end_column=9)
    return ws


def sheet_findings(wb):
    ws = wb.create_sheet("Findings")
    ws.sheet_properties.tabColor = "833C00"
    titleblock(ws, "Findings register (doc 40)",
               "Stable IDs — never renumbered. Any change to a sizing figure or architecture decision "
               "must be reflected here with its ID.", 6)
    cols = [("ID", 9), ("Track", 11), ("Finding", 46), ("Consequence", 74), ("Status", 30), ("Lives in", 22)]
    header(ws, cols, row=4)
    r = 5
    for fid, track, one, cons, status, where in FINDINGS:
        put(ws, r, 1, fid, "s_lbl"); put(ws, r, 2, track); put(ws, r, 3, one)
        put(ws, r, 4, cons); put(ws, r, 5, status); put(ws, r, 6, where)
        # Match on substring, not prefix: four rows qualify their status rather than
        # leading with the keyword ("Corrosion rate PENDING M4", "design intent;
        # PENDING H/G", "settled (safety register 2)", "binding requirement"), and a
        # prefix test dropped all four through unstyled. Status TEXT is unchanged —
        # only the derived fill. REQUIRED is tested first: "REQUIRED-PENDING J"
        # contains both and must read as required, not merely pending.
        if "REQUIRED" in status:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=BADF)
            ws.cell(row=r, column=5).font = Font(name="Calibri", size=10, bold=True, color=BADT)
        elif "PENDING" in status:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=WARN)
        elif status.startswith("settled"):
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=OKG)
        elif "requirement" in status:
            # Same treatment the Register sheet gives grade "requirement".
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=OKG)
            ws.cell(row=r, column=5).font = Font(name="Calibri", size=10, bold=True, color=OKT)
        else:
            raise ValueError(
                f"FINDINGS row {fid}: unrecognised status {status!r}. Expected one "
                "containing 'REQUIRED', containing 'PENDING', beginning 'settled', or "
                "containing 'requirement' — an unmatched status renders unstyled and "
                "reads as a further, undefined state.")
        r += 1
    ws.auto_filter.ref = f"A4:F{r-1}"
    ws.freeze_panes = "C5"
    r += 1
    band(ws, r, "CORRECTION-TRAIL PRINCIPLE (doc 40 §5)", 6); r += 1
    put(ws, r, 1, "Every erratum stays visible. Never cite a margin as a computed value; re-run every "
                  "comfort claim when the design point moves; solve steady states simultaneously, never "
                  "as hand chains. The correction trail is part of the design's credibility — and, for a "
                  "defensive publication, part of the evidentiary record.", "s_txt")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=6); r += 3
    band(ws, r, "THE LIQUID-TRACK ERRATA TRAIL (doc 12 §2)", 6); r += 1
    for i, (e, lesson) in enumerate([
        ("Regeneration heat overstated 2× in first-pass tables", "Never cite a margin as a computed value"),
        ("Absorber outlet is a band, not a point; cooling reclassified optional → REQUIRED", ""),
        ("Diffuser dynamic wet pressure was missing from the column blower budget", "Drilled rings win"),
        ("Nylon removed from approved materials", "CaCl₂ stress-cracks loaded polyamides"),
        ("Crystallization threshold corrected: 40 wt% liquidus ~12–13 °C, not ~5", "Interlock restated at 42/43 wt%"),
        ("Regen exhaust condensate: 53 °C dew point rains salty condensate", "Sloped drip-leg routing mandatory"),
        ("Sparger head constant 1.35–1.4 kPa per 10 cm at SG 1.4", ""),
        ("Once-through could not deliver 4-adult comfort at DP-A (59–66% RH)", "Re-run every comfort claim when the design point moves"),
        ("The self-consistent steady state, not the hand chain, sizes the system", "Hand chains drifted 15–30% on duty"),
    ], start=1):
        put(ws, r, 1, f"Erratum {i}", "s_lbl"); put(ws, r, 2, e)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        put(ws, r, 5, lesson, "s_note")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        r += 1
    return ws


def sheet_safety(wb):
    ws = wb.create_sheet("Safety & Materials")
    ws.sheet_properties.tabColor = "C00000"
    titleblock(ws, "Safety register & materials law",
               "BINDING. Any build omitting these departs from this design (doc 00 §8, doc 11 §1).", 4)
    header(ws, [("#", 6), ("Requirement", 34), ("Specification", 96), ("Source", 18)], row=4)
    r = 5
    for n, req, spec, src in SAFETY:
        put(ws, r, 1, n, "s_num", "0"); put(ws, r, 2, req, "s_lbl")
        put(ws, r, 3, spec); put(ws, r, 4, src)
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFF2F2")
        r += 1
    r += 1
    band(ws, r, "THE TWO-WORLDS MATERIALS LAW — everything brine- or raw-water-wetted is metal-free", 4); r += 1
    hdr = r
    for i, t in enumerate(["Status", "Materials", "Rationale", ""], start=1):
        put(ws, hdr, i, t, "s_hdr")
    r += 1
    for status, mats, why in MATERIALS:
        put(ws, r, 1, status, "s_lbl"); put(ws, r, 2, mats); put(ws, r, 3, why)
        if status == "Never" or status.startswith("Prohibited"):
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BADF)
        elif status == "Use freely":
            ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=OKG)
        r += 1
    r += 1
    put(ws, r, 1, "Rule of thumb: parts sold for pools, aquaculture, or chemical dosing are probably right; "
                  "parts sold for plumbing hide brass. Titanium heat exchangers are the ONLY membrane between "
                  "the two worlds — never let the fluids mix. These rules are never relaxed for land installs, "
                  "because the desiccant itself is the chloride source.", "s_txt")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=4)
    return ws


def sheet_checks(wb):
    ws = wb.create_sheet("Checks")
    ws.sheet_properties.tabColor = "FFC000"
    titleblock(ws, "Consistency observations",
               "Raised while compiling this register by re-deriving each published figure. "
               "These are observations for the authors, not errata — no document has been changed.", 7)
    cols = [("ID", 9), ("Type", 20), ("Where", 18), ("Observation", 52),
            ("What the arithmetic shows", 74), ("Impact", 26), ("Disposition", 62)]
    header(ws, cols, row=4)
    r = 5
    for row in CHECKS:
        for i, v in enumerate(row, start=1):
            put(ws, r, i, v, "s_lbl" if i == 1 else "s_txt")
        imp = row[5]
        fill = OKG if imp.startswith("Informational") else (WARN if imp.startswith("Medium") else GREY)
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=fill)
        r += 1
    ws.auto_filter.ref = f"A4:G{r-1}"
    r += 1
    put(ws, r, 1, "Method: each headline figure was re-derived independently from the stated basis using the "
                  "Magnus formulation and the doc 00 §4 model, then compared with the published value. "
                  "CHK-006 through CHK-008 record figures that reproduced exactly and are logged as "
                  "confirmations, not defects. Per the repository's own doctrine (doc 50 §8), nothing here "
                  "should be silently corrected — supersede with a version bump and leave the trail visible.", "s_note")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=7)
    return ws


def sheet_sources(wb):
    ws = wb.create_sheet("Sources")
    titleblock(ws, "Source document set", "The lineage this register was compiled from.", 4)
    header(ws, [("Document", 44), ("Version", 12), ("Contents", 96), ("", 4)], row=4)
    r = 5
    for doc, ver, contents in DOCS:
        put(ws, r, 1, doc, "s_lbl"); put(ws, r, 2, ver); put(ws, r, 3, contents)
        r += 1
    r += 1
    band(ws, r, "SCOPE & PROVENANCE", 4); r += 1
    for k, v in [
        ("Reading order", "00 (shared basis) → 10–12 (liquid) → 20–22 (solid) → 30–31 (integration, upgrades) "
                          "→ 40 (findings) → 50 (disclosure)"),
        ("Single source rule", "Doc 00 is the single source for scope, design point, shared psychrometrics, the "
                               "airflow–moisture model, the CO₂ stack, the X8 doctrine, and the safety register. "
                               "Track docs reference it, never restate it"),
        ("What this workbook is", "A derived working artifact. The documents remain authoritative; where this "
                                  "workbook and a document disagree, the document wins and the Checks sheet "
                                  "records the discrepancy"),
        ("What this workbook is not", "Not a measurement record. No figure here is measured data unless its "
                                      "grade says so — and none currently does"),
        ("Regeneration", "python3 scripts/build_parameter_workbook.py [output.xlsx] — regenerate rather than "
                         "hand-edit, so the register stays traceable to the documents"),
        ("Licensing", "Workbook content CC-BY-4.0 with the documentation; generator script MIT with scripts/"),
        ("Disclosure", f"Open defensive publication · concept DOI {CONCEPT_DOI} · version DOI {VERSION_DOI} · "
                       "no patents sought or held"),
    ]:
        put(ws, r, 1, k, "s_lbl"); put(ws, r, 2, v)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    return ws


# ================================================================== MAIN =====
def main(out):
    wb = Workbook()
    wb.remove(wb.active)
    styles(wb)
    sheet_cover(wb)
    sheet_legend(wb)
    sheet_designpoint(wb)
    sheet_register(wb)
    sheet_psychro(wb)
    sheet_model(wb)
    sheet_liquid(wb)
    sheet_solid(wb)
    sheet_co2(wb)
    sheet_heatwater(wb)
    sheet_upgrades(wb)
    sheet_tests(wb)
    sheet_findings(wb)
    sheet_safety(wb)
    sheet_checks(wb)
    sheet_sources(wb)

    wb.properties.title = "Heat-Driven Comfort & Water — Parameter Register"
    wb.properties.subject = "Defensive publication parameter register at DP-A"
    wb.properties.creator = "slice-cooling · scripts/build_parameter_workbook.py"
    wb.properties.description = (
        "Every quantitative claim in the slice-cooling document lineage, with confidence grades, "
        "gating tests, sources, and live re-derivations. Paper design — nothing built.")
    wb.properties.keywords = "liquid desiccant; aluminium fumarate; DP-A; defensive publication; prior art"
    wb.calculation.fullCalcOnLoad = True

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(out)
    print(f"wrote {out}  ({len(wb.worksheets)} sheets, {len(REGISTER)} register rows, "
          f"{len(TESTS)} tests, {len(FINDINGS)} findings)")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "parameter_register.xlsx")
